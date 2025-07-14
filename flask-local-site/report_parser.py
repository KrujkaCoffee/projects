import os
import stat
import pickle
from time import sleep
import shutil
from openpyxl import Workbook, load_workbook


import project_cust_38.Cust_SQLite as CSQ
import project_cust_38.Cust_Functions as F


from html2image import Html2Image

import settings


def get_pictures_type(is_html=False):
    pictures_di = {}
    pictures_li = []
    png_id = 0
    for pic in os.listdir(settings.PICTURE_FOLDER):
        if pic.endswith('.png'):
            png_id += 1
            one_picture_dict = {}
            title = pic[:-4]
            if is_html:
                pictures_di[png_id] = f'{title}.html'

            else:
                one_picture_dict['path'] = os.path.join(settings.STATIC_PATH, pic)
                one_picture_dict['title'] = title
                one_picture_dict['png_id'] = png_id
                pictures_li.append(one_picture_dict)
    if is_html:
        return pictures_di
    else:
        return pictures_li


def save_png(full_path):
    with open(full_path, encoding='utf-8') as file:
        response = file.read()
    hti = Html2Image(size=(settings.WIDTH, settings.HEIGHT))
    hti.output_path = settings.PICTURE_FOLDER
    _, filename = os.path.split(full_path)
    title = filename[:-5]
    hti.screenshot(response, save_as=f'{title}.png')


def check_pic_foler():
    if not os.path.exists(settings.PICTURE_FOLDER):
        os.mkdir(settings.PICTURE_FOLDER)
        os.chmod(settings.PICTURE_FOLDER, stat.S_IRUSR | stat.S_IWUSR | stat.S_IXUSR)

def get_pictures_and_html():
    check_pic_foler()
    try:
        for filename in os.listdir(settings.REPORT_FOLDER):
            full_path = os.path.join(settings.REPORT_FOLDER, filename)
            if os.path.isfile(full_path) and full_path.endswith('.html'):
                save_png(full_path)
                shutil.copy(full_path, settings.FULL_TEMP_HTML_PATH)
    except FileNotFoundError:
        ...

def save_stat(create_time_files):
    if os.path.exists(settings.STAT_FILE):
        os.remove(settings.STAT_FILE)
    with open(settings.STAT_FILE, 'wb') as f:
        pickle.dump(create_time_files, f)

def create_stats():
    create_time_files = {}
    try:
        for file in os.listdir(settings.REPORT_FOLDER):
            if file.endswith('.html'):
                file_path = os.path.join(settings.REPORT_FOLDER, file)
                create_time_files[file_path] = os.stat(file_path).st_ctime
    except FileNotFoundError:
        ...
    return create_time_files

def load_stat():
    if not os.path.exists(settings.STAT_FILE):
        return {}
    with open(settings.STAT_FILE, 'rb') as f:
        create_time_files = pickle.load(f)
    return create_time_files

def clean_old_files():
    for file in os.listdir(settings.PICTURE_FOLDER):
        full_path = os.path.join(settings.PICTURE_FOLDER, file)
        if full_path.endswith('.png'):
            os.remove(full_path)

    for html_file in os.listdir(settings.FULL_TEMP_HTML_PATH):
        full_path = os.path.join(settings.FULL_TEMP_HTML_PATH, html_file)
        os.remove(full_path)


def check_stat():
    old_stats = load_stat()
    new_stats = create_stats()
    if not old_stats == new_stats:
        print('пересобирание картинок из html')
        clean_old_files()
        get_pictures_and_html()
        save_stat(new_stats)
    else:
        print('html без изменений действий не требуется')


def start_daemon_thread(interval):
    print(f'проверка html будет провоедиться через {settings.TIME_CHECK_HTML} минут')
    count_check = 0
    while True:
        count_check += 1
        print(f'check # {count_check}')
        check_stat()
        sleep(interval * 60)


# def dict_etapi( conn = '',cur = ''):
#     self.DICT_ETAPI = dict()
#     custom_request_c = f'''SELECT * FROM operacii'''
#     SPIS_OP = CSQ.custom_request_c(settings.NARYAD_DB,custom_request_c,hat_c=False, conn=conn, cur = cur,rez_dict=True)
#     if SPIS_OP == False:
#         return False
#     for i in range(len(SPIS_OP)):
#         self.DICT_ETAPI[SPIS_OP[i]['name']] = SPIS_OP[i]['etap']
    
def load_res(nom_mk:int, conn = '',cur= ''):
    query = f'''SELECT data FROM res WHERE Номер_мк == {nom_mk}
            '''
    res = CSQ.custom_request_c(settings.RES_XML, query, conn=conn, cur = cur)
    try:
        if res[-1][0] == '':
            return False
    except:
        print(f'ошибка загрузки мк {nom_mk}')
        return False
    return F.from_binary_pickle(res[-1][0])

def dict_opers():
    custom_request_c = f'''SELECT * FROM operacii'''
    SPIS_OP = CSQ.custom_request_c(settings.NARYAD_DB, custom_request_c,hat_c=False,rez_dict=True)
    di = F.deploy_dict_c(SPIS_OP,'name')
    return di

def podrazdel_kod():
    res = []
    custom_request_c = f'''SELECT rab_c.Код, rab_c.Имя FROM rab_c'''
    spis_cexov = CSQ.custom_request_c(settings.USERS_DB, custom_request_c, hat_c=False)
    for cex in spis_cexov:
        res.append((cex[0], cex[1]))
    return res




DICT_OPER_FULL = dict_opers()
DEPATMENTS = podrazdel_kod()


def check_report(report, podrazd, nach=False, konec=False, ):
    if report == 'Реестр проектов в работе':
        res = ready_procent_ver2(nach, konec, podrazd)
    return res





def ready_procent_ver2(nach, konec, podrazd='010301', *args):
    query = f'''SELECT naryad.Номер_мк, naryad.Подтвержд_вып_дата 
    , mk.Номер_заказа 
    , mk.Номер_проекта 
    , mk.Вид 
    , mk.Вес 
    , naryad.ФИО 
    , naryad.Фвремя 
    , naryad.ФИО2 
    , naryad.Фвремя2 
    , naryad.Твремя 
    , naryad.Операции 
    , naryad.Опер_время 
    FROM naryad INNER JOIN mk ON mk.Пномер = naryad.Номер_мк WHERE mk.Статус == "Открыта" '''
    resp = CSQ.custom_request_c(r'C:\DB_srv\Naryad.db', query,rez_dict=True)
    dict_mk = dict()
    for item in resp:
        list_opers = item['Операции'].split("|")
        list_time = item['Опер_время'].split("|")
        if item['Номер_мк'] not in dict_mk:
            dict_mk[item['Номер_мк']] = {
                'Номер_заказа':item['Номер_заказа'],
                'Номер_проекта': item['Номер_проекта'],
                'Вид': item['Вид'],
                'Вес': item['Вес'],
                'План': dict(),
                'Факт': dict(),
            }
        for i, oper in enumerate( list_opers):
            oper_name= oper.split("$")[-1]
            if oper_name not in DICT_OPER_FULL:
                continue
            etap = DICT_OPER_FULL[oper_name]['rc']
            time = list_time[i]
            if etap not in dict_mk[item['Номер_мк']]['План']:
                dict_mk[item['Номер_мк']]['План'][etap] = 0
                dict_mk[item['Номер_мк']]['Факт'][etap] = 0
            dict_mk[item['Номер_мк']]['Факт'][etap] += F.valm(time)
    
    
    for nom_mk in dict_mk:
        res = load_res(nom_mk)
        for dse in res:
            for oper in dse['Операции']:
                kod = oper['Опер_РЦ_код']
                time = oper['Опер_Тпз'] +  oper['Опер_Тшт']
                if kod not in dict_mk[nom_mk]['План']:
                    dict_mk[nom_mk]['План'][kod] = 0
                    dict_mk[nom_mk]['Факт'][kod] = 0
                dict_mk[nom_mk]['План'][kod] += time
    
        

    res = [['Номер_МК','Номер_заказа','Номер_проекта','Вид','Вес,кг*(с поправкой)','Всего','Освоено',"Процент %",'Вес_осталось,кг']]
    for nom_mk in dict_mk:
        plan = 0
        fact = 0
        for kod in dict_mk[nom_mk]['План']:
            if kod[:4] == podrazd[:4]:
                plan+=dict_mk[nom_mk]['План'][kod]
        for kod in dict_mk[nom_mk]['Факт']:
            if kod[:4] == podrazd[:4]:
                fact += dict_mk[nom_mk]['Факт'][kod]
        if fact <=0:
            continue
        proc = 0
        if plan >0:
            proc = round(fact/plan*100,1)
        ves = round(F.valm(dict_mk[nom_mk]['Вес'])/1.482,1)
        ost=round((100-proc)*ves/100,1)
        if ost < 0:
            ost = 0
        tmp= [nom_mk,dict_mk[nom_mk]['Номер_заказа'],dict_mk[nom_mk]['Номер_проекта'],dict_mk[nom_mk]['Вид']
            ,ves,round(plan/480,1),round(fact/480,1),proc,ost]
        res.append(tmp)

    res= F.sort_by_column_c(res,'Процент %')
    res.append(['' for _ in res[0]])
    res.append(['' for _ in res[0]])
    return res


def excel_maker(full_li, name):
    wb = Workbook(write_only=False, iso_dates=False)
    ws = wb.active
    ws.title = 'title'
    for row_num, row in enumerate(full_li,1):
        for col_num, value in enumerate(row,1):
            ws.cell(row=col_num, column=row_num, value=value)
    wb.save(name)

















if __name__ == '__main__':
    # p = Thread(target=start_daemon_thread, args=(settings.TIME_CHECK_HTML,), daemon=True)
    # p.start()
    # sleep(80)
    check_stat()