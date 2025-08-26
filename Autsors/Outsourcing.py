import copy
import os
import pathlib
import sys
import re
from collections import defaultdict
from enum import Enum

import requests
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import Qt, QSettings
from PyQt5.QtGui import QColor, QCursor
from PyQt5.QtWidgets import QMainWindow
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
import win32com.client as win32

from mydesign3 import Ui_MainWindow
from project_cust_38 import Cust_Qt as CQT
from project_cust_38 import Cust_Functions as F
from project_cust_38 import Cust_mes as CMS
from project_cust_38 import Cust_SQLite as CSQ
from project_cust_38 import Cust_docs as CDOCS
from project_cust_38 import Cust_config as CFG
import order as OERP
import tk_operation_docs as TKDOCS

import user_filters as UF
import mini_window as MINWIN


settings = QSettings('config.ini', QSettings.IniFormat)


def dict_from_combo_table(tbl: QtWidgets.QTableWidget) -> list[dict]:
    result = []
    headers = [
        tbl.horizontalHeaderItem(col).text()
        for col in range(tbl.columnCount())
    ]
    for row in range(tbl.rowCount()):
        tmp = {}
        for col, head in enumerate(headers):
            cell = tbl.cellWidget(row, col)
            if isinstance(cell, QtWidgets.QComboBox):
                text = cell.currentText()
            else:
                text = tbl.item(row, col).text()
            tmp[head] = text
        result.append(tmp)
    return result

def dict_from_info_table(tbl: QtWidgets.QTableWidget) -> dict:
    result = {}
    for row in range(tbl.rowCount()):
        key = tbl.item(row, 0).text()
        cell = tbl.cellWidget(row, 1)
        if isinstance(cell, QtWidgets.QComboBox):
            val = cell.currentText()
        else:
            val = tbl.item(row, 1).text()
        result[key] = val
    return result

def aggregate_by_key(data: list[dict], key):
    result = {}
    for dic in data:
        if key in dic:
            result.setdefault(dic[key], []).append(dic)
    return result

def extract_one_value(data):
    if isinstance(data, (list, tuple)) and len(data) == 1:
        return extract_one_value(data[0])
    return data


def num_row_key_by_info_table(tbl: QtWidgets.QTableWidget, key: str) -> int | None:
    for row in range(tbl.rowCount()):
        if tbl.item(row, 0).text() == key:
            return row

def insert_application(order: dict, naryads: list[dict], order_params, nar_params):
    order_fields = F.deploy_dict_c(order_params, 'Поле')
    nar_fields = F.deploy_dict_c(nar_params, 'Поле')

    query = '''INSERT INTO outsource(user, create_date) VALUES(?, ?) RETURNING id'''
    data = CSQ.custom_request_c(F.scfg('Naryad'), query, list_of_lists_c=[F.user_full_namre(), F.now()],  one=True, one_column=True, hat_c=False)
    if pk := extract_one_value(data):
        result_order = result_nar = None
        list_of_list = []
        insert_order_param = f'''INSERT INTO out_params(param_id, value, order_id) VALUES(?, ?, ?)'''
        for key, value in order.items():
            param = order_fields.get(key)
            if param_id := param.get('Пномер'):
                list_of_list.append([param_id, value, pk])
        result_order = CSQ.custom_request_c(F.scfg('Naryad'), insert_order_param, list_of_lists_c=list_of_list)
        insert_nar_param = f'''INSERT INTO out_params(param_id, value, order_id, Номер_наряда) VALUES(?, ?, ?, ?)'''
        list_of_list = []
        for naryad in naryads:
            for key, value in naryad.items():
                param = nar_fields.get(key)
                if param:
                    param_id = param.get('Пномер')
                    nar_id = naryad.get('Номер_наряда')
                    list_of_list.append([param_id, value, pk, nar_id])
        if len(list_of_list) > 0:
            result_nar = CSQ.custom_request_c(F.scfg('Naryad'), insert_nar_param, list_of_lists_c=list_of_list)
        if pk and result_nar and result_order:
            CQT.msgbox('Заявка успешно создана!')
        else:
            query_drop = (
                f'''DELETE FROM outsource WHERE id = {pk}''',
                f'''DELETE FROM out_params WHERE order_id = {pk}'''
            )
            result = all(CSQ.custom_request_c(F.scfg('Naryad'), q) for q in query_drop)
            CQT.msgbox('Что-то пошло не так.. попробуйте снова')
            if not result:
                print(f'Не удалось удалить неудачно созданную заявку: №{pk}')

def get_applications():
    active = settings.value('order.search.active', defaultValue=False, type=bool)
    query = f"""
    select знпр.№ERP, знпр.№проекта, mk.НомКплан, mk.Пномер as №MK,
            o.НаименованиеДокумента as НаименованиеДокумента,
        o.id as "Номер_заявки", 
        o.create_date as "Создано", 
        o.user as "Инициатор",
        o.НаУдаление,
        "" as "Docs" 
    from outsource o
    left join out_params op on op.order_id = o.id
    inner join naryad n on n.Пномер = op.Номер_наряда
    inner join mk on mk.Пномер = n.Номер_мк
    inner join пл_оуп по on по.НомПл = mk.НомКплан
    inner join знпр on знпр.s_num = по.Пномер_ЗП
    where op.Номер_наряда != '' and o.НаУдаление != {int(active)}
    group by o.id
    """
    apps = CSQ.custom_request_c(F.scfg('Naryad'), query, rez_dict=True, attach_dbs=F.scfg('DB_kplan'))
    return apps

def select_naryads(Номер_заявки: int, obj_name: str = 'Наряд'):
    stmt = f"""
        SELECT naryad.*, GROUP_CONCAT(op.value , '|') AS param_values, GROUP_CONCAT(opi.Поле , '|') AS param_fields
        FROM naryad
        LEFT JOIN out_params_info opi ON opi.Объект = "{obj_name}"
        LEFT JOIN out_params op ON opi.Пномер = op.param_id AND op.order_id = {Номер_заявки} AND op.Номер_наряда = naryad.Пномер
        WHERE op.Номер_наряда = naryad.Пномер AND op.order_id = {Номер_заявки}
        GROUP BY naryad.Пномер
    """
    res = CSQ.custom_request_c(F.scfg('Naryad'), stmt, rez_dict=True)
    result = []
    for row in res:
        tmp = {}
        if fields := row.get('param_fields'):
            fields = fields.split('|')
            values = row.get('param_values', '')
            if values is None or len(fields) != len(values.split('|')):
                values = [''] * len(fields)
            else:
                values = values.split('|')
            tmp.update(zip(fields, values))
            tmp.update(row)
        result.append(tmp)
    if isinstance(result, list):
        return result

def none_to_empty_string(data: list[dict]) -> list[dict]:
    response = []
    for row in data:
        tmp = row.copy()
        for key, value in row.items():
            tmp[key] = '' if value is None else value
        response.append(tmp)
    return response

def get_operations(Операции: list[str], Номенклатурный_номер: str, ОперациПоНаряду: str):
    data = eval(Операции)
    unpacked_opers = ОперациПоНаряду.split('|')
    result = []
    for oper in data:
        for target_oper in unpacked_opers:
            num, name = target_oper.split('$')
            if oper['Опер_номер'] == num:
                result.append(oper)
    return result

def get_resources(Номер_мк, ДСЕ_ID, Операции):
    # result = CSQ.custom_request_c(F.scfg('db_resxml'),
    #                                  f'''SELECT data FROM res WHERE Номер_мк = {Номер_мк!r}''',
    #                                  rez_dict=True, one=True)
    # if not isinstance(result, dict):
    #     CQT.msgbox('Не удалось загрузить ресурсную по МК: %s' % Номер_мк)
    #     return result
    ids = ДСЕ_ID.split('|')
    oper_ids = Операции.split('|')
    if len(ids) != len(oper_ids):
        return []
    data = []
    result = CMS.load_res(Номер_мк)
    by_dse_id = defaultdict(dict)
    for item in result:
        cp_item = item.copy()
        opers = []
        for dse_pk, oper in zip(ids, oper_ids):
            if dse_pk == str(item['Номерпп']):
                # opers.append(oper)
                for key, value in item.items():
                    cp_item[str(key)] = str(value)
                by_dse_id[dse_pk]['ДСЕ'] = cp_item
                by_dse_id[dse_pk].setdefault('Операции', list()).append(oper)

                # cp_item['ОперациПоНаряду'] = '|'.join(opers)
                # data.append(cp_item)
    for pk, item in by_dse_id.items():
        dse_elem = item['ДСЕ']
        dse_elem['ОперациПоНаряду'] = '|'.join(item['Операции'])
        data.append(dse_elem)
    return data

def get_wo_autsorce():
    stmt = """
        SELECT naryad.*
        FROM naryad 
        LEFT JOIN jurnal on jurnal.Номер_наряда=naryad.Пномер 
        INNER JOIN mk ON mk.Пномер = naryad.Номер_мк 
        LEFT JOIN out_params op ON naryad.Пномер = op.Номер_наряда 
        WHERE Аутсорсинг=1 and naryad.Пномер NOT IN (select distinct Номер_наряда
from out_params op
left join outsource o on o.id = op.order_id 
where o.НаУдаление != 1 and Номер_наряда != '')
        GROUP BY naryad.Пномер
    """
    res = CSQ.custom_request_c(F.scfg('Naryad'), stmt, rez_dict=True)
    if isinstance(res, list):
        return res

class  ObjBranch:
    def __init__(self, name, parent, pk: str, param_id: str, child_id: str, view_format: str, get_data, decor_table = None, **kwargs):
        self.name = name
        self.pk = pk
        self.child_id = child_id
        self.param_id = param_id
        self.fn = get_data
        self.parent: ObjBranch = parent
        self.child: ObjBranch = None
        self.view_format = view_format
        self.current = None
        self.data = None
        self.decor_table = decor_table
        self.__dict__.update(kwargs)


    @property
    def current_id(self):
        if not isinstance(self.current, dict): return
        if self.pk in self.current:
            return int(self.current[self.pk])

    @property
    def fields(self):
        fields = settings.value(f'{self.name}.fields', None)
        if fields is None and isinstance(self.current, dict):
            fields = {key: True for key in self.current.keys() if key != self.pk}
            self.fields = fields
        return fields

    @fields.setter
    def fields(self, new_val: dict[str: bool]):
        settings.setValue(f'{self.name}.fields', new_val)

    def get_data(self):
        kwargs = {}
        if self.parent and self.parent.child_id and self.parent.current:
            for pk in self.parent.child_id:
                kwargs[pk] = self.parent.current.get(pk)
        return self.fn(**kwargs)

    def __str__(self):
        if isinstance(self.current, dict):
            return self.view_format.format(**self.current)
        return ''


class TreeObjs:
    OBJECTS = {
        'Заявка': {
            'view_format': 'Заявка № {Номер_заявки}',
            'get_data': get_applications,
            'decor_table': '',
            'child_id': ('Номер_заявки',),
            'pk': 'Номер_заявки',
            'param_id': 'order_id'
        },
        'Наряд': {
            'view_format': 'Наряд № {Пномер}',
            'get_data': select_naryads,
            'child_id': ('Номер_мк', 'ДСЕ_ID', 'Операции'),
            'pk': 'Пномер',
            'param_id': 'Номер_наряда'
        },
        'ДСЕ': {
            'view_format': '{Наименование} {Номенклатурный_номер}',
            'get_data': get_resources,
            'pk': 'Номерпп',
            'child_id': ('Операции', 'Номенклатурный_номер', 'ОперациПоНаряду'),
            'param_id': 'dse_id'
        },
        'Операция': {
            'view_format': '{Опер_наименование}',
            'get_data': get_operations,
            'pk': 'Опер_код',
            'child_id': ('Материалы',),
            'param_id': 'oper_id'
        },
        'Материал': {
            'view_format': '{Мат_наименование}',
            'get_data': lambda Материалы: eval(Материалы) if isinstance(Материалы, str) else [],
            'pk': '',
            'child_id': '',
            'param_id': ''
        }
    }

    __inst = {}
    def __init__(self):
        prev = None
        for obj, kwargs in self.OBJECTS.items():
            inst = ObjBranch(name=obj, parent=prev, **kwargs)
            if prev:
                prev.child = inst
            self.__inst[obj] = inst
            prev = inst

    def __getitem__(self, item) -> ObjBranch:
        if item in self.__inst:
            return self.__inst[item]

    def __iter__(self):
        return iter(self.__inst)

    def __enter__(self):
        self.context = []

    def first_object(self) -> ObjBranch | None:
        for _, val in self.__inst.items():
            return val
        return None


class Utils:
    def __init__(self):
        self.mk_data = F.scfg('mk_data')

    def open_tk_path(self, mk_id: str | int, tk_name: str, nn: str) -> list[list[str]]:
        mk_id = str(mk_id)
        path = pathlib.Path(F.scfg('mk_data')) / mk_id / f'{tk_name}_{nn}.pickle'
        if path.exists():
            return F.open_file_c(str(path), False, "|", pickl=True)

    def unpack_oper_docs(self, tk: list[list[str]]):
        result = []
        for oper in tk[10:]:
            if len(oper) >= 15 and oper[15]:
                result.extend(oper[15].split('%20'))
        return result

    @staticmethod
    def replace_none_with_an_empty_string(data: list[dict]):
        result = []
        for idx, item in enumerate(data):
            cp_item = item.copy()
            for key, value in item.items():
                if value is None:
                    value = ''
                cp_item[key] = value
                result.append(cp_item)
        return result


class ObjParam:
    def __init__(self, *, pk: int, dict_podrazdel: dict, param_type: str, department: str, default_val: str = '', sep: str = '|'):
        self.pk = pk
        self.sep = sep
        self.DICT_PODRAZDEL = dict_podrazdel
        self.type = param_type
        self.default_val = self.make_default_params(default_val)
        self.department = department.split(sep)
        self.r, self.g, self.b = self.rgb_by_department()

    def make_default_params(self, default_val: str):
        default_val = str() if default_val is None else default_val
        match self.type:
            case 'enum':
                return default_val.split(self.sep)
            case 'bool':
                return  ['не выбрано', 'да', 'нет']
            case _:
                return default_val

    def rgb_by_department(self):
        resp_department = self.department[0] if len(self.department) >= 1 else ''
        return self.DICT_PODRAZDEL.get(resp_department, '255;255;255').split(';')


def get_object_params(*, Объект: str, order_id, Номер_наряда = None, dse_id = None, oper_id = None):
    kwargs = locals().copy()
    obj_name = kwargs.pop('Объект')
    if not any(kwargs.values()):
        return []
    # where = ' AND '.join(f'({key} = {value!r} OR {key} IS NULL)' for key, value in kwargs.items() if value is not None)
    fields = ' AND '.join(f'{key} = {value!r}' for key, value in kwargs.items() if value is not None)
    where =  f'AND {fields}' if fields else ''
    query = f"""
        SELECT op.Пномер, op.value, opi.Подразделение, opi.Поле, opi.Описание, opi.По_умолчанию
        FROM out_params_info opi
        LEFT JOIN out_params op ON opi.Пномер = op.param_id {where}
        WHERE Объект = {obj_name!r}
    """
    response = CSQ.custom_request_c(CFG.Config.project.db_naryad, query, rez_dict=True)
    if isinstance(response, list):
        return Utils.replace_none_with_an_empty_string(response)


def contains_uuid(s):
    if not isinstance(s, str):
        return s
    if s.strip() == "":
        return False
    uuid_pattern = r'\b[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}\b'
    if re.search(uuid_pattern, s):
        return False
    return True

def refresh_main_tbl_after_do(fn):
    def wrap(self, *args, **kwargs):
        last_object_state = self.current_obj_row
        result = fn(self, *args, **kwargs)
        self.set_objs(last_object_state)
        return result
    return wrap

def change_logger(action: str, data: str):
    fio = F.user_full_namre()
    query = "INSERT INTO out_journal(Событие, ФИО, Данные) VALUES(?, ?, ?)"
    return CSQ.custom_request_c(CFG.Config.project.db_naryad, query, list_of_lists_c=[[action, fio, data]])

class BtnAddLineState(Enum):
    CHANGE_TYPE = 'Сменить тип документа в ERP'
    EDIT_DOC = 'Редактировать документ'
    CREATE_DOC = 'Создать документ в ERP'


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.versia = '1.5.4'
        self.NAME_MODULE_BASE = 'Аутсорсинг'
        self.DICT_PARTNERS = {}
        self.db_naryad = F.scfg('Naryad')
        self.db_kplan = F.scfg('DB_kplan')
        self.db_dse = F.scfg('BD_dse')
        self.mk_data = F.scfg('mk_data')
        self.settings = QtCore.QSettings('config.ini', QSettings.IniFormat)
        self.utils = Utils()
        search_by_order = self.settings.value('chk_filtr_by_order.checked', True, type=bool)
        self.ui.chk_filtr_by_order.setChecked(search_by_order)
        deps = CSQ.custom_request_c(self.db_kplan, 'SELECT Наименование_ЕРП, Цвет FROM podrazdel', rez_dict=True)
        self.DICT_PODRAZDEL = F.deploy_dict_c(deps, 'Наименование_ЕРП') if deps else {}
        status = CSQ.custom_request_c(self.db_kplan, 'SELECT Имя, color FROM status_poz', rez_dict=True)
        self.DICT_STATUS_POZ = F.deploy_dict_c(deps, 'Имя') if status else {}
        CMS.dict_emploee_full(
            bd_users=CFG.Config.project.db_users,
            self=self
        )
        self.erp_client = OERP.OrderERP(erp_base_name='ERP_MES2', window=self)
        order_services = self.erp_client.get_order_services()
        self.ERP_ORDER_SERVICES = F.deploy_dict_c(order_services, 'Description')


        self.ui.tabWidget.currentChanged.connect(self.tab_click)
        # self.ui.tbl_infozayav.itemSelectionChanged.connect(self.info_selected)

        self.ui.btn_ok_edit_z.clicked.connect(self.apply_edit)
        self.ui.btn_cansel_edit_z.clicked.connect(lambda *_: self.fill_new_zayav())
        self.ui.btn_apply_chosed.clicked.connect(self.create_order)
        # self.ui.tbl_list_zayv.doubleClicked.connect(lambda *_: self.ui.tabWidget.setCurrentIndex(1))
        self.ui.btn_reload_list_zayav.clicked.connect(lambda *_: self.set_objs(0))
        self.ui.btn_new_zayav.clicked.connect(lambda *_: self.click_new_app())
        self.ui.btn_chose_nar.clicked.connect(self.choose_naryad)

        self.ui.btn_showhide_tree.clicked.connect(lambda *_: self.ui.tree_objs.setHidden(not self.ui.tree_objs.isHidden()))

        self.current_table = 'outsource'
        self.current_oper = None

        self.department = self.get_current_department()
        self.department = 'Планово-диспетчерский отдел Производства'
        self.order_status = [
            'Создание',
            # 'Указание документов, сроков и материалов',
            'Подготовка',
            'Поиск поставщика',
            'Согласование',
            'Размещение заказа поставщику',
            'Приемка ТО',
            'Приемка ОТК',
            'Доработка',
            'Закрытие'
        ]
        base_struct = {'data': {}, 'current': None}
        self.obj_headers = ['Категория', 'Объект', 'Фильтр', 'Поля', 'Сброс', 'Отображать в заголовках']
        self.obj_keys = ['Заявка', 'Наряд', 'ДСЕ', 'Операция', 'Материал']

        self.current_info_naryad = None
        self.info_focus_table = None
        self.current_focus_tbl = None
        self.tree_objs2 = TreeObjs()
        self.tree_objs = {
            key: base_struct.copy()
            for key in self.obj_keys
        }
        self.obj_current_row = 0
        self.current_main_row = 0
        self.opacity_ratio = 255/6

        params = CSQ.custom_request_c(F.scfg('Naryad'), '''SELECT * FROM out_params_info''', rez_dict=True)
        self.info_params: [str, ObjParam] = self.make_params_obj(params)

        self.DICT_PARAMS_BY_OBJECT = aggregate_by_key(params, 'Объект')

        self.fill_tree_obj_table()

        self.current_item = None
        checked = settings.value('order.search.active', defaultValue=False, type=bool)
        self.ui.chk_filtr_active_only.setChecked(checked)

        # menu_actions
        self.menu_excel_action = QtWidgets.QAction('Выгрузить EXCEL')
        self.ui.menu.addAction(self.menu_excel_action)

        # on_item_selected
        self.ui.tbl_list_zayv.itemSelectionChanged.connect(self.order_selected)
        self.ui.tbL_addit_params.itemSelectionChanged.connect(self.on_select_param)

        # on_clicked
        self.ui.tree_objs.clicked.connect(lambda m_idx: self.set_objs(m_idx.row()))
        self.ui.tbl_list_zayv.clicked.connect(self.item_zayav_selected)
        self.ui.tbl_new_order.clicked.connect(self.on_select_new_order)
        self.ui.tbl_list_chosed_naryads.clicked.connect(self.on_select_choosed_naryad)
        self.ui.chk_filtr_active_only.clicked.connect(self.fill_active_zayav)
        self.ui.btn_add_param.clicked.connect(self.add_param_template)
        self.ui.btn_del_param.clicked.connect(self.drop_param_template)
        self.ui.tbl_infozayav.clicked.connect(self.on_clicked_info)
        self.ui.tbl_list_vals_params.clicked.connect(self.focus_tbl)
        self.ui.tbL_addit_params.clicked.connect(self.focus_tbl)
        self.ui.btn_add_line.clicked.connect(self.btn_create_or_change_order_type)
        self.ui.btn_delete_line.clicked.connect(self.mark_remove_order_in_ERP)

        self.ui.cld.clicked.connect(self.calendar_selected)
        self.ui.btn_clear_chosed.clicked.connect(self.clear_selected_naryads)

        # on_double_clicked
        self.ui.tbl_list_zayv.doubleClicked.connect(self.set_objs)

        # on_text_changed
        self.ui.cmb_select_obj_level.currentTextChanged.connect(self.fill_params_table)
        self.ui.tbl_list_vals_params.itemChanged.connect(self.on_text_param_changed)

        # on_state_changed
        self.ui.chk_filtr_by_order.stateChanged.connect(self.normalize_filter)

        # on_triggered
        self.menu_excel_action.triggered.connect(self.dump_excel)
        # set_cursor
        self.ui.tree_objs.verticalHeader().setStyleSheet("QHeaderView::section { cursor: pointer; }")
        self.set_objs(0)
        self.user_filters = UF.UserFilter(
            settings=self.settings,
            tbl=self.ui.tbl_list_zayv,
            filtr=self.ui.tbl_list_zayv_filtr,
            combo=self.ui.cmb_name_filtr,
            label=self.ui.le_name_filtr,
            btn_save=self.ui.btn_save_name_filtr,
            name_filter='outsource_list_zayv_user_filter'
        )
        CFG.Config.user_config.load_user_config(self)
        CQT.load_icons(self,24)
        CQT.connect_to_resize(self, CMS.tmp_dir())
        CMS.add_action_config_save_tbl_filtrs(self, self.ui)

    def make_params_obj(self, params):
        return {
            param.get('Поле'): ObjParam(
                pk=param.get('Пномер'),
                dict_podrazdel=self.DICT_PODRAZDEL,
                param_type=param.get('Тип'),
                department=param.get('Подразделение', ''),
                default_val=param.get('По_умолчанию', ''),
            )
            for param in params
        }

    def get_attached_docs(self, current_doc):

        from project_cust_38 import api_erp_commands as AEC
        order = self.tree_objs2.first_object()
        if not order:
            return
        item = order.current
        if not item:
            return CQT.msgbox('Не выбрана заявка')

        prefix, doc_name = current_doc.split('_', 1)

        query = f"""
         ВЫБРАТЬ
         	{doc_name}ПрисоединенныеФайлы.Наименование КАК Наименование,
         	{doc_name}ПрисоединенныеФайлы.Размер КАК Размер,
         	{doc_name}ПрисоединенныеФайлы.Расширение КАК Расширение
        ИЗ
            Справочник.{doc_name}ПрисоединенныеФайлы КАК {doc_name}ПрисоединенныеФайлы
        ГДЕ {doc_name}ПрисоединенныеФайлы.ВладелецФайла = &ВладелецФайла И {doc_name}ПрисоединенныеФайлы.ПометкаУдаления = ЛОЖЬ
        """
        order_id = item.get('Номер_заявки')
        order = CSQ.custom_request_c(
            CFG.Config.project.db_naryad,
            f"SELECT Ref_Key, НаименованиеДокумента FROM outsource WHERE id = {order_id}",
            rez_dict=True,
            one=True
        )
        if not order:
            return
        ref = order['Ref_Key']
        if ref is None:
            return
        refs = AEC.Refs_wet(query)
        ref_obj = AEC.Ref_wet('ВладелецФайла', 'Документы.ЗаказПереработчику2_5', ref)
        refs.add_ref(ref_obj)
        code, data = AEC.get_wet_request(text=query, refs=refs)
        files = data['data']

        if files is None:
            return CQT.msgbox('Произошла ошибка при попытке запроса прикрепленных файлов 1С')
        if files:
            return [f'{file["Наименование"]}.{file["Расширение"]}' for file in files]
        return []


    def get_docs(self, Номер_заявки, doc_name, *args):
        oper_docs = TKDOCS.OperationDocs(self, None)
        # tbl = self.ui.tbl_list_zayv
        # cur_data = CQT.get_dict_line_form_tbl(tbl)
        nars = select_naryads(Номер_заявки=Номер_заявки)
        dse = []
        result = [['ID', 'ДСЕ', 'Операция', 'Документ', 'Просмотреть', 'Прикрепить']]
        poki = CFG.Config.place.poki
        attached_docs = self.get_attached_docs(doc_name)
        for nar in nars:
            dse_names = nar['ДСЕ'].split('|')
            nom_mk = nar['Номер_мк']
            dse.extend((item.split('$')[1], nom_mk) for item in dse_names)

        for (nn, nom_mk) in dse:
            dse_item = CSQ.custom_request_c(CFG.Config.project.db_dse,
                                            f'select * from dse where Номенклатурный_номер = {nn!r} and poki = {poki}',
                                            rez_dict=True, one=True)
            tk = dse_item['Номер_техкарты']
            name = f'{tk}_{nn}.pickle'
            lst_tk = F.open_file_c(F.scfg("mk_data") + os.sep + str(nom_mk) + os.sep + name, False, '|', pickl=True,
                                   propuski=True)[11:]
            for oper in lst_tk:
                if len(oper) >= 15 and oper[15]:
                    for doc in oper[15].split('%20'):
                        file_id, form, name, full_name = oper_docs.unpack_include_format(doc)
                        if file_id:
                            text_btn = '✅ Прикреплено'
                            if full_name not in attached_docs:
                                text_btn = '<attach_check>|Прикрепить'
                            result.append([file_id, nn, oper[0], full_name, '<view_btn>|👁', text_btn])
        return result

    def oform_docs_table(self, tbl: QtWidgets.QTableWidget):
        nk_attach = CQT.num_col_by_name_c(tbl, 'Прикрепить')
        nk_id = CQT.num_col_by_name_c(tbl, 'ID')
        uploaded_docs = []
        for row_idx in range(tbl.rowCount()):
            text = 'Открепить' if tbl.item(row_idx, nk_id).text() in uploaded_docs else 'Прикрепить'
            CQT.add_btn(tbl, row_idx, nk_attach, text, conn_func_checked_row_col=self.attach_tflex_doc_to_1c, self=tbl)

    def attach_tflex_doc_to_1c(self, tbl: QtWidgets.QTableWidget, row, *args):
        doc_id = CQT.valt(tbl, 'ID', row)
        doc_name = CQT.valt(tbl, 'Документ', row)
        if F.is_numeric(doc_id):
            client = CDOCS.TFlexFileManager()
            content = client.get_doc_by_id(doc_id)
            print(content)
            #todo прикрепить документ к заказу переработчика 1с


    def generate_lst_mk_path(self, mk_nums):
        gen_unpack_nn = lambda temp: (dse.split('$')[1] for dse in temp.split('|'))
        docs = []
        for mk in mk_nums:
            dse = mk.get('ДСЕ')
            mk_id = mk.get('Номер_мк')
            for nn in gen_unpack_nn(dse):
                nom_tk = CSQ.custom_request_c(self.db_dse,f'SELECT Номер_техкарты FROM dse WHERE Номенклатурный_номер = {nn!r}', one=True, one_column=True, hat_c=False)
                if isinstance(nom_tk, list) and len(nom_tk) == 1:
                    tk = self.utils.open_tk_path(mk_id, nom_tk[0], nn)
                    if tk_docs := self.utils.unpack_oper_docs(tk):
                        docs.extend(tk_docs)
        return set(docs)

    def add_param_template(self, *args):
        self.info_focus_table = 'param'
        tbl: QtWidgets.QTableWidget = self.current_focus_tbl
        nk_pk = CQT.num_col_by_name_c(tbl, 'Пномер')
        if nk_pk is not None:
            for row in range(tbl.rowCount()):
                pk = tbl.item(row, nk_pk).text()
                if pk.strip() == '':
                    return
        tbl_param = self.ui.tbL_addit_params
        tbl_val = self.ui.tbl_list_vals_params
        if tbl and tbl in (tbl_param, tbl_val):
            empty_row = ['' for _ in range(tbl.columnCount())]
            old_values = CQT.list_from_wtabl_c(tbl, hat_c=True)
            old_values.append(empty_row)
            CQT.fill_wtabl(old_values, tbl, set_editeble_col_nomera={0}, StretchLastSection=False)

    def focus_tbl(self, index, *args):
        self.current_focus_tbl = self.sender()

    def drop_param_template(self, *args):
        tbl = self.ui.tbL_addit_params
        current_row = CQT.get_dict_line_form_tbl(self.ui.tbL_addit_params)
        if pk := current_row.get('Пномер'):
            param_name = current_row.get('Поле')
            result = CSQ.custom_request_c(self.db_naryad,
                                          f'SELECT COUNT(*) FROM out_params WHERE param_id = {pk}',
                                          hat_c=False, one_column=True, one=True)
            if result == False:
                return CQT.msgbox('Произошел сбой в бд')
            if result[0] == 0:
                if CQT.msgboxgYN('Вы уверены что хотите удалить параметр: "%s"' % param_name):
                    CSQ.custom_request_c(self.db_naryad, f'DELETE FROM out_params_info WHERE Пномер = {pk}')
                    CQT.msgbox('Параметр успешно удален')
            else:
                CQT.msgbox('Невозможно удалить параметр т.к. он уже используется')
            self.fill_params_table()
        else:
            tbl.removeRow(tbl.currentRow())

    def normalize_filter(self, new_state: int):
        self.settings.setValue('chk_filtr_by_order.checked', new_state)

    def get_current_department(self):
        fio = F.user_full_namre()
        query = f'''SELECT Подразделение FROM employee WHERE ФИО = {fio!r} '''
        response = CSQ.custom_request_c(F.scfg('BD_users'), query, one_column=True, one=True, hat_c=False)
        if dep := extract_one_value(response):
            return dep
        CQT.msgbox('Не удалось вычислить ваше подразделение')

    def fill_tree_obj_table(self):
        data = [self.obj_headers]
        for idx, obj in enumerate(self.obj_keys):
            symbol = '- ' if idx == 0 else '└ '
            template = [''] * len(self.obj_headers)
            template[0] = ('  ' * idx) + symbol + obj
            data.append(template)

        CQT.fill_wtabl(data, self.ui.tree_objs, height_row=22, head_column=0, min_width_col=140, StretchLastSection=True)

        nk_fields = CQT.num_col_by_name_c(self.ui.tree_objs, 'Поля')
        nk_reset = CQT.num_col_by_name_c(self.ui.tree_objs, 'Сброс')
        nk_view_head = CQT.num_col_by_name_c(self.ui.tree_objs, 'Отображать в заголовках')
        for row in range(self.ui.tree_objs.rowCount()):
            obj_name = self.obj_keys[row]
            cur_obj = self.tree_objs2[obj_name]

            cell = MINWIN.CustomComboBox(cur_obj, on_checked=self.on_checked_hidden_column)
            self.ui.tree_objs.setCellWidget(row, nk_fields, cell)
            CQT.add_btn(self.ui.tree_objs,
                        row,
                        nk_reset,
                        text='▼',
                        self=cur_obj.name,
                        conn_func_checked_row_col=self.on_clicked_clear_filters
                        )
            CQT.add_check_box(self.ui.tree_objs, row, nk_view_head, conn_func_checked_row_col=self.on_checked_view_in_headers)

    def get_object_params(self, obj_name: str):
        kwargs = {
            # 'Объект': obj_name,
            'order_id': 'Заявка',
            'Номер_наряда': 'Наряд',
            'dse_id': 'ДСЕ',
            'oper_id': 'Операция',
        }
        result_kwargs = {
            'Объект': obj_name,
            'order_id': None,
            'dse_id': None,
            'oper_id': None,
            'Номер_наряда': None
        }
        lst_objects = list(TreeObjs.OBJECTS)
        rank = lst_objects.index(obj_name)
        rng = lst_objects[:rank + 1]
        for pk_name, cr_obj_name in kwargs.items():
            value = None
            if cr_obj_name in rng:
                value = self.tree_objs2[cr_obj_name].current_id
            result_kwargs[pk_name] = value
        # order_id = self.tree_objs2['Заявка'].current_id
        # Номер_наряда = self.tree_objs2['Наряд'].current_id
        # dse_id = self.tree_objs2['ДСЕ'].current_id
        # oper_id = self.tree_objs2['Операция'].current_id
        return get_object_params(
            **result_kwargs
            # Объект=obj_name,
            # **kwargs
            # order_id=order_id,
            # dse_id=dse_id,
            # oper_id=oper_id,
            # Номер_наряда=Номер_наряда
        )

    def inject_params_into_object_item(self, obj_item, data):
        cp_item = data.copy()
        obj_item.current = cp_item
        params = self.get_object_params(obj_item.name)
        new_dic = {field.get('Поле'):  field.get('value') for field in params}
        new_dic.update(cp_item)
        return new_dic

    def get_data_object(self, name):
        tbl = self.ui.tree_objs
        cur_obj = self.tree_objs2[name]
        data = cur_obj.get_data()
        result = []
        for item in data:
            result.append(
                self.inject_params_into_object_item(cur_obj, item)
            )

        nk_view_head = CQT.num_col_by_name_c(tbl, 'Отображать в заголовках')
        for obj_idx, obj_name in enumerate(self.obj_keys[:self.current_obj_row]):
            tbl = self.ui.tree_objs
            widget = tbl.cellWidget(obj_idx, nk_view_head)
            if isinstance(widget, QtWidgets.QCheckBox):
                head_data = self.tree_objs2[obj_name].current
                if widget.isChecked():
                    result = [{**head_data, **elem} for elem in result]
        return result

    def on_checked_hidden_column(self) -> None:
        self.set_objs(current_row=self.current_obj_row)

    def on_checked_view_in_headers(self, checked: bool, row: int, *args) -> None:
        data = self.get_data_object(self.current_obj_name)
        if data and isinstance(data, list) and any(bool(item) for item in data):
            fields = self.tree_objs2[self.current_obj_name].fields
            if isinstance(fields, dict) and len(fields) > 0:
                data = self.filtr_fields(data)
            self.fill_tbl_list_zayv(data)
        else:
            CQT.msgbox('Вложенных данных в объекте не найдено')
            self.current_obj_row -= 1

    def on_clicked_clear_filters(self, cur_obj: str, *args) -> None:
        settings.setValue(f'{cur_obj}.fields', None)
        self.set_objs(current_row=self.current_obj_row)

    def on_clicked_info(self, index, *args):
        tbl = self.ui.tbl_infozayav
        tbl_vals = self.ui.tbl_list_vals_params
        match self.info_focus_table:
            case 'param':
                if tbl.item(index.row(), 0).text() == 'По_умолчанию':
                    cur_default = tbl.item(index.row(), 1).text()
                    CQT.fill_wtabl(cur_default.split('|'), tbl_vals, hide_head_rows=True, hide_head_column=True, set_editeble_col_nomera={0})


    def order_selected(self, *args):
        tbl_info = self.ui.tbl_infozayav
        tbl_info.setRowCount(0)
        data_for_table = []
        self.info_focus_table = 'main'
        data = defaultdict(dict)
        tbl_val = CQT.get_dict_line_form_tbl(self.ui.tbl_list_zayv)
        new_item = self.inject_params_into_object_item(self.tree_objs2[self.current_obj_name], tbl_val)
        self.tree_objs2[self.current_obj_name].current = new_item

        self.change_create_order_btn_state()
        if tbl_val.get('НаУдаление') == '1':
            self.ui.btn_delete_line.setText('Снять с удаления')
            self.ui.btn_add_line.setEnabled(False)
        else:
            self.ui.btn_delete_line.setText('Поставить удаление')
            self.ui.btn_add_line.setEnabled(True)

        for param, value in tbl_val.items():
            param_object = self.info_params.get(param)
            if param_object:
                department = param_object.department[0]
                data[department][param] = value
            else:
                data['Информационные'][param] = value

        for category, params in data.items():
            data_for_table.append([category, '!span!'])
            for param, value in params.items():
                data_for_table.append([param, value])
        CQT.fill_wtabl(data_for_table, tbl_info, hide_head_column=True, hide_head_rows=True, height_row=22, ogr_maxshir_kol=110)
        self.unpack_combobox_on_table(tbl_info)
        self.unpack_erp_params(tbl_info)


    def fill_active_zayav(self, checked: bool, *args):
        settings.setValue('order.search.active', checked)
        self.set_objs(0)

    def on_select_choosed_naryad(self, *args):
        self.info_focus_table = 'naryad'
        tbl_info = self.ui.tbl_infozayav
        self.current_info_naryad = self.ui.tbl_list_chosed_naryads.currentRow()
        data = [
            [key, value]
            for key, value in CQT.get_dict_line_form_tbl(self.ui.tbl_list_chosed_naryads).items()
        ]
        CQT.fill_wtabl(data, tbl_info, hide_head_column=True, hide_head_rows=True, height_row=22)
        self.unpack_combobox_on_table(tbl_info)

    @property
    def current_obj_row(self):
        count_obj = len(self.obj_keys)
        if self.obj_current_row < 0:
            self.obj_current_row = 0
        if self.obj_current_row >= count_obj:
            self.obj_current_row = count_obj - 1
        return self.obj_current_row
    
    @property
    def current_obj_name(self):
        return self.obj_keys[self.current_obj_row]

    @current_obj_row.setter
    def current_obj_row(self, val: int):
        val = 0 if val < 0 else val if val < len(self.obj_keys) else len(self.obj_keys) - 1
        self.obj_current_row = val
        self.ui.tree_objs.selectRow(val)
        self.set_color_tree_obj_row()

    def on_select_new_order(self, *args):
        self.info_focus_table = 'order'
        tbl_info = self.ui.tbl_infozayav
        data = [
            [key, value]
            for key, value in CQT.get_dict_line_form_tbl(self.ui.tbl_new_order).items()
        ]
        CQT.fill_wtabl(data, tbl_info, hide_head_column=True, hide_head_rows=True, height_row=22)
        self.unpack_combobox_on_table(tbl_info)

    def on_text_param_changed(self, item: QtWidgets.QTableWidgetItem, *args):
        tbl_info = self.ui.tbl_infozayav
        tbl = self.ui.tbl_list_vals_params
        data = CQT.list_from_wtabl_c(tbl)
        new_state = '|'.join(item[0] for item in data if item[0].strip())
        row = num_row_key_by_info_table(self.ui.tbl_infozayav, 'По_умолчанию')
        row and tbl_info.item(row, 1).setText(new_state)

    def on_select_param(self, *args):
        self.info_focus_table = 'param'
        tbl = self.ui.tbl_infozayav
        enums = {
            'Подразделение': list(self.DICT_PODRAZDEL.keys()),
            'Статус': self.order_status,
            'Тип': ['Строка', 'Перечисление', 'да/нет', 'Число'],
            'Объект': self.obj_keys
        }
        dict_val = CQT.get_dict_line_form_tbl(self.ui.tbL_addit_params)
        data = []
        for key, val in dict_val.items():
            data.append([key, val])

        CQT.fill_wtabl(data, tbl, hide_head_rows=True, hide_head_column=True, set_editeble_col_nomera={1})
        for row, (key, val) in enumerate(data):
            if key == 'По_умолчанию':
                CQT.fill_wtabl(val.split('|'), self.ui.tbl_list_vals_params, hide_head_rows=True, hide_head_column=True,
                               set_editeble_col_nomera={0})
            item = tbl.item(row, 1)
            if key == 'Пномер':
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if key == 'Описание':
                tbl.setRowHeight(row, 80)
                tbl.item(row, 1).setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
            if val := enums.get(key):
                cell = QtWidgets.QComboBox()
                cell.addItems(val)
                tbl.setCellWidget(row, 1, cell)
            tbl.setItem(row, 1, item)

    def create_unicode_icon(self, unicode_char):
        # Создаем QPixmap из текста Unicode
        font = QtGui.QFont("Arial", 32)  # Устанавливаем шрифт и размер
        pixmap = QtGui.QPixmap(32, 32)  # Размер иконки
        pixmap.fill(Qt.transparent)  # Заполняем прозрачным цветом

        painter = QtGui.QPainter(pixmap)
        painter.setFont(font)
        painter.drawText(pixmap.rect(), Qt.AlignCenter, unicode_char)
        painter.end()
        return QtGui.QIcon(pixmap)

    def unpack_combobox_on_table(self, tbl):
        for row in range(tbl.rowCount()):
            key_item = tbl.item(row, 0)
            key = key_item.text()
            key_item.setToolTip(key)
            item = tbl.item(row, 1)
            value = item.text()
            if value == '!span!':
                tbl.setSpan(row, 0, 1, 2)
                span = tbl.item(row, 0)
                span.setTextAlignment(Qt.AlignCenter)
                r, g, b = self.DICT_PODRAZDEL.get(key, '255;255;255').split(';')
                span.setBackground(QtGui.QColor(int(r), int(g), int(b), int(255 / 3)))
                font = span.font()
                font.setBold(True)
                span.setFont(font)

            if param_obj := self.info_params.get(key):
                if self.department in param_obj.department:
                    if param_obj.type == 'date':
                        item.setIcon(self.create_unicode_icon("📅"))

                    if isinstance(param_obj.default_val, list):
                        CQT.add_combobox(self, tbl, row, 1, param_obj.default_val, first_void=False)
                        tbl.cellWidget(row, 1).setStyleSheet(
                            "QComboBox { background-color: rgba(%s); }" % f'{255},{255},{255}')
                        item.setText('')
                    else:
                        value = item.text() if item.text().strip() else param_obj.default_val
                        item.setText(value)
                        # item.setFlags(item.flags() | Qt.ItemIsEditable)
                CQT.set_color_wtab_c(tbl, row, 0, int(param_obj.r), int(param_obj.g), int(param_obj.b), int(255 / 4))
                # CQT.set_color_wtab_c(tbl, row, 1, int(param_obj.r), int(param_obj.g), int(param_obj.b), int(255 / 4))
                CQT.set_color_wtab_c(tbl, row, 1, 255, 255, 255, int(255 / 4))


    def unpack_erp_params(self, tbl):
        for row in range(tbl.rowCount()):
            key = tbl.item(row, 0) and tbl.item(row, 0).text()
            item = tbl.item(row, 1)
            param_obj = self.info_params.get(key)
            services_for_choice = [''] + list(self.ERP_ORDER_SERVICES.keys())
            if key == 'Услуга':
                item.setText(item.text())
                CQT.add_combobox(self, tbl, row, 1, services_for_choice, first_void=False)
                tbl.cellWidget(row, 1).setStyleSheet(
                    "QComboBox { background-color: rgba(%s); }" % f'{param_obj.r},{param_obj.g},{param_obj.b},{255 / 4}')
                item.setText('')

    def fill_tbl_list_zayv(self, data: list):
        CQT.fill_wtabl(data, self.ui.tbl_list_zayv)
        CQT.fill_filtr_c(self, self.ui.tbl_list_zayv_filtr, self.ui.tbl_list_zayv, hidden_scroll=True)
        self.set_color_departments(self.ui.tbl_list_zayv)
        self.set_color_tree_obj_row()
        self.decor_table()

    def set_color_tree_obj_row(self):
        tbl = self.ui.tree_objs
        for row in range(tbl.rowCount()):
            r, g, b = 255, 255, 255
            if row == self.current_obj_row:
                r, g, b = 188, 220, 244
            CQT.set_color_row_wtab_c(tbl, row, r, g, b)
            tbl.verticalHeaderItem(row).setBackground(QColor(r, g, b))

    def get_parent(self):
        tbl = self.ui.tree_objs
        current_row = tbl.currentRow()
        lst_values = CQT.list_from_wtabl_c(tbl, rez_dict=True, hat_c=False)
        parent_row = 0 if current_row <= 0 else current_row - 1
        return lst_values[parent_row]

    def select_row(self):
        tbl = self.ui.tbl_list_zayv
        selected_items = tbl.selectedItems()
        if selected_items:
            selected_row = selected_items[0].row()
            if selected_row > 0:
                for i in range(tbl.rowCount()):
                    tbl.verticalHeaderItem(i).setBackground(None)
                tbl.verticalHeaderItem(selected_row - 1).setBackground(Qt.yellow)

    def set_objs(self, current_row: int = 1, *args):
        tbl = self.ui.tree_objs
        if self.ui.tbl_list_zayv.hasFocus():
            self.current_obj_row += 1
            current_row = self.obj_current_row
        if self.ui.tree_objs.hasFocus():
            self.current_obj_row = tbl.currentRow()
            current_row = self.ui.tree_objs.currentRow()
        if current_row >= len(self.obj_keys):
            current_row = len(self.obj_keys) - 1
        name = self.obj_keys[current_row]
        nk_obj = CQT.num_col_by_name_c(tbl, 'Объект')
        nk_filtr = CQT.num_col_by_name_c(tbl, 'Фильтр')
        for row in range(current_row + 1, tbl.rowCount()):
            tbl.item(row, nk_obj).setText('')
            tbl.item(row, nk_filtr).setText('')
        parent = self.tree_objs2[name].parent
        if parent and parent.current is None:
            CQT.msgbox(f'Сначала необходимо выбрать родительский объект {parent.name}')
            return
        data = self.get_data_object(name)
        if data == [] and current_row == 0:
            CQT.clear_tbl(self.ui.tbl_infozayav)
            return CQT.clear_tbl(self.ui.tbl_list_zayv)
        if data and isinstance(data, list) and any(bool(item) for item in data):
            data = self.filtr_fields(data)
            self.fill_tbl_list_zayv(data)
            CQT.clear_tbl(self.ui.tbl_infozayav)
            for obj_row in range(self.current_obj_row + 1):
                name = self.obj_keys[obj_row]
                instance = self.tree_objs2[name]
                item_obj = tbl.item(obj_row, nk_obj)
                item_filtr = tbl.item(obj_row, nk_filtr)
                item_obj and item_obj.setText(str(instance))
                item_filtr and item_filtr.setText(', '.join(instance.child_id))
        else:
            self.current_obj_row -= 1


    def get_list_of_dict_headers(self, list_of_dicts: list[dict]):
        if list_of_dicts and isinstance(list_of_dicts, list) and isinstance(list_of_dicts[0], dict):
            return list(list_of_dicts[0].keys())

    def filtr_fields(self, data):
        tree_obj = self.tree_objs2[self.current_obj_name]
        if tree_obj.fields is None:
            if new_fields := self.get_list_of_dict_headers(data):
                tree_obj.fields = dict.fromkeys(new_fields, True)
            else:
                return
        cp_data = copy.deepcopy(data)
        for idx, row in enumerate(cp_data):
            for head, value in row.items():
                if head in tree_obj.fields and not tree_obj.fields[head]:
                    data[idx].pop(head)
        return data

    def set_color_departments(self, tbl: QtWidgets.QTableWidget):
        headers = [tbl.horizontalHeaderItem(c).text() for c in range(tbl.columnCount())]
        for row in range(tbl.rowCount()):
            for col, head in enumerate(headers):
                if info_params_inst := self.info_params.get(head):
                    text = (
                            'Ответствтенноые за занесение подразделения: \n%s' %
                            '\n'.join(f'* {department}' for department in info_params_inst.department)
                    )
                    tbl.item(row, col).setToolTip(text)
                    CQT.set_color_wtab_c(tbl, row, col,
                                         info_params_inst.r,
                                         info_params_inst.g,
                                         info_params_inst.b,
                                         int(self.opacity_ratio))

    def get_order_ref_key(self):
        tbl = self.ui.tbl_list_zayv
        btn = self.ui.btn_add_line

        # current_row = CQT.get_dict_line_form_tbl(tbl)
        current_order = self.tree_objs2.first_object()
        if not current_order:
            return
        current_row = current_order.current
        order_id = current_row.get('Номер_заявки')
        delete_mark = current_row.get('НаУдаление', '0')
        btn.setEnabled(str(delete_mark) == '0')
        if order_id and delete_mark == '0':
            # btn.setEnabled(True)
            order_id = int(order_id)
            response = CSQ.custom_request_c(
                CFG.Config.project.db_naryad,
                f'SELECT Ref_Key FROM outsource WHERE id = {order_id}',
                rez_dict=True,
                one=True
            )
            if isinstance(response, dict):
                ref_key = response.get('Ref_Key')
                if isinstance(ref_key, str) and len(ref_key) >= 36:
                    return ref_key

    def change_create_order_btn_state(self, *args):
        ref_key = self.get_order_ref_key()
        btn = self.ui.btn_add_line
        current_order = self.tree_objs2.first_object()
        if not current_order: return
        order_data = current_order.current
        calced_doc_type = self.get_doc_type_by_order_info()
        current_doc_type = order_data.get('НаименованиеДокумента')
        if current_doc_type and calced_doc_type != current_doc_type:
            text = BtnAddLineState.CHANGE_TYPE.value
        elif ref_key:
            text = BtnAddLineState.EDIT_DOC.value
        else:
            text = BtnAddLineState.CREATE_DOC.value
        btn.setText(text)

    def item_zayav_selected(self, model_index: QtCore.QModelIndex, *args):
        tbl = self.ui.tree_objs
        tbl_order = self.ui.tbl_list_zayv
        row = model_index.row()
        col = model_index.column()
        nk_type_doc = CQT.num_col_by_name_c(tbl_order, 'НаименованиеДокумента')
        type_doc = ref_key = None
        if self.current_obj_row == 0:
            nk_order_id = CQT.num_col_by_name_c(tbl_order, 'Номер_заявки')
            order_id = tbl_order.item(row, nk_order_id).text()
            response = CSQ.custom_request_c(
                CFG.Config.project.db_naryad,
                f'SELECT Ref_Key, НаименованиеДокумента FROM outsource WHERE id = {order_id}',
                rez_dict=True,
                one=True
            )
            ref_key = type_doc = None
            if isinstance(response, dict):
                ref_key = response.get('Ref_Key')
                type_doc = response.get('НаименованиеДокумента')

        if col == nk_type_doc and type_doc and ref_key:
            code, data = self.erp_client.get_document(type_doc, ref_key)
            if code == 200:
                result = [
                    [key, value]
                    for key, value in data.items()
                    if contains_uuid(value) and not isinstance(value, (list, dict))
                ]
                CQT.fill_wtabl(result, self.ui.tbl_infozayav, hide_head_rows=True, hide_head_column=True)
        else:
            tbl_values = CQT.get_dict_line_form_tbl(self.ui.tbl_list_zayv)
            self.current_main_row = self.ui.tbl_list_zayv.currentRow()
            current_row = self.obj_current_row
            if current_row >= len(self.obj_keys):
                current_row = len(self.obj_keys) - 1
            name = self.obj_keys[current_row]
            if nk_obj := CQT.num_col_by_name_c(tbl, 'Объект'):
                if inst := self.tree_objs2[name]:
                    inst.current = tbl_values
                    self.change_create_order_btn_state()

                nk_filtr = CQT.num_col_by_name_c(tbl, 'Фильтр')
                item_obj = tbl.item(self.obj_current_row, nk_obj)
                item_filtr = tbl.item(self.obj_current_row, nk_filtr)
                item_obj and item_obj.setText(str(inst))
                item_filtr and item_filtr.setText(', '.join(inst.child_id))

    def add_search_fields(self, field):
        tbl = self.ui.tree_objs
        nk_fields = CQT.num_col_by_name_c(tbl, 'Поля')
        item = tbl.item(self.obj_current_row, nk_fields)
        if item is not None:
            old_val = item.text()
            new_val = '|'.join((old_val, field))
            tbl.item(self.obj_current_row, nk_fields).setText(new_val)

    def clear_fields(self):
        tbl = self.ui.tree_objs
        nk_fields = CQT.num_col_by_name_c(tbl, 'Поля')
        item = tbl.item(self.obj_current_row, nk_fields)
        if item is not None:
            tbl.item(self.obj_current_row, nk_fields).setText('')

    def apply_filter(self, key_val):
        filters = {
            self.ui.tbl_list_zayv_filtr: self.ui.tbl_list_zayv,
            self.ui.tbl_list_naryads_filtr: self.ui.tbl_list_naryads
        }
        for filter, tbl in filters.items():
            if filter.hasFocus() and key_val == 16777220:
                CMS.apply_filtr_c(self, filter, tbl)

    def keyReleaseEvent(self, a0, QKeyEvent=None):
        key = a0.key()
        self.apply_filter(key)
        text = self.ui.tabWidget.tabText(self.ui.tabWidget.currentIndex())
        if key == 16777268 and text == 'Перечень':
            self.set_objs(0)
        if self.ui.cmb_name_filtr.hasFocus():
            if key == QtCore.Qt.Key_Delete:
                self.user_filters.del_filt_pl_user_filtrs()

    def structure_selected(self, info_table: QtWidgets.QTableWidget, tbl_filter: QtWidgets.QTableWidget):
        self.ui.tbl_infozayav.setRowCount(0) or self.ui.tbl_infozayav.setColumnCount(0)
        data = CQT.get_dict_line_form_tbl(info_table)
        CQT.fill_wtabl(
            [[key, val] for key, val in data.items()],
            self.ui.tbl_infozayav,
            set_editeble_col_nomera={1},
            hide_head_column=True,
            hide_head_rows=True,
        )
        CQT.fill_filtr_c(self, tbl_filter, info_table, hidden_scroll=True)

    def check_nar_reg_for_order(self, pks: list[int]):
        pks_to_str = f'= {pks[0]}' if len(pks) == 1 else f'IN {tuple(pks)}'
        query = f'SELECT Номер_наряда FROM out_params WHERE Номер_наряда {pks_to_str}'
        pks = set(str(pk) for pk in CSQ.custom_request_c(self.db_naryad, query, one_column=True, hat_c=False))
        if pks == set():
            return True
        return CQT.msgbox('На наряды: {pks} уже зарегистрированы заявки'.format(pks=', '.join(pks)))

    @CQT.onerror
    def create_order(self, *args):
        if not CMS.user_access(self.db_naryad, 'аутсорс_создание_заявки', F.user_name()):
            return
        tbl_order = self.ui.tbl_new_order
        tbl_naryads = self.ui.tbl_list_chosed_naryads
        order = dict_from_combo_table(tbl_order)
        naryads = CQT.list_from_wtabl_c(tbl_naryads, hat_c=False, rez_dict=True)

        order_params = self.DICT_PARAMS_BY_OBJECT.get('Заявка')
        nar_params = self.DICT_PARAMS_BY_OBJECT.get('Наряд')
        if self.check_nar_reg_for_order([nar['Номер_наряда'] for nar in naryads]):
            app = insert_application(order=order.pop(0), naryads=naryads, order_params=order_params,
                                     nar_params=nar_params)
            if app:
                naryads = ','.join(str(naryad['Номер_наряда']) for naryad in naryads)
                user = F.user_full_namre()
                query = f"""SELECT 
                    naryad.Пномер as Номер_Наряда,
                    mk.Пномер as МК, 
                    пл_оуп.НомПл as "Номер КПЛ", 
                    знпр.№ERP, 
                    знпр.№проекта
                     FROM naryad
                            JOIN mk ON mk.Пномер = naryad.Номер_мк
                            join пл_оуп ON пл_оуп.НомПл = mk.НомКплан
                            JOIN знпр ON пл_оуп.Пномер_ЗП = знпр.s_num
                            where naryad.Пномер IN ({naryads})
                        """
                credentials = CSQ.custom_request_c(CFG.Config.project.db_naryad, query, rez_dict=True,
                                                   attach_dbs=self.db_kplan)
                msg = f"{user} Создал(а) заявку на аутсорсинг по нарядам:\n"
                for nar_credentials in credentials:
                    msg += "\n".join(
                        f"\t\t{key}: {val}" for key, val in nar_credentials.items()
                    )
                    msg += '\n' + ('===' * 16) + '\n'
                CMS.send_info_mk_b24_by_action(msg, 'Аутсорсинг')
                print()
                CQT.msgbox('Заявка успешно создана!')
                self.ui.tabWidget.setCurrentIndex(0)

    def info_selected(self):
        tbl = self.ui.tbl_infozayav
        col_item = tbl.item(tbl.currentRow(), 0)
        if col_item is not None:
            col_name = col_item.text()
            cur_obj = self.obj_keys[self.obj_current_row]
            param = self.get_param_line_by_field(cur_obj, col_name)
            if param:
                self.ui.fr_cld.setVisible('date' == param.get('Тип'))
        else:
            self.ui.fr_cld.setVisible(False)

    def calendar_selected(self, date: QtCore.QDate):
        tbl = self.ui.tbl_infozayav
        item_key = tbl.item(tbl.currentRow(), 0)
        if item_key is not None:
            param_obj = self.info_params.get(item_key.text())
            if param_obj and param_obj.type == 'date':
                formatted_date = date.toString("yyyy-MM-dd")
                item = tbl.item(tbl.currentRow(), 1)
                if item is None: return
                item.setText(formatted_date)

    def prepare_info_table(self, info_table, tbl_filter):
        info_table.setSelectionBehavior(QtWidgets.QTableWidget.SelectRows)
        info_table.itemSelectionChanged.connect(lambda: self.structure_selected(info_table, tbl_filter))

    def fill_params_table(self, curr_obj = None):
        if curr_obj is None:
            curr_obj = self.ui.cmb_select_obj_level.currentText()
        tbl = self.ui.tbL_addit_params
        response = CSQ.custom_request_c(F.scfg('Naryad'), f'SELECT * FROM out_params_info WHERE Объект = {curr_obj!r}', rez_dict=True)
        if isinstance(response, list):
            CQT.fill_wtabl(none_to_empty_string(response), tbl)

    def tab_click(self, index):
        self.ui.tbl_infozayav.setRowCount(0)
        self.ui.tbl_infozayav.setColumnCount(0)
        text = self.ui.tabWidget.tabText(index)
        match text:
            case 'Перечень':
                self.set_objs(0)
            case 'Новая заявка':
                self.ui.groupBox.setVisible(True)
                self.click_new_app()
            case 'Управление параметрами':
                self.ui.cmb_select_obj_level.clear()
                self.ui.cmb_select_obj_level.addItems(self.obj_keys)
                self.fill_params_table()

    def get_doc_type_by_order_info(self):
        tbl = self.ui.tbl_list_zayv
        key_use_mats = 'Использование сырья'
        current_order = self.tree_objs2.first_object()
        cur_order_row = current_order.current
        match cur_order_row.get(key_use_mats):
            case 'Давальческое':
                return 'Document_ЗаказПереработчику2_5'
            case 'Поставщика':
                return 'Document_ЗаказПоставщику'

    @refresh_main_tbl_after_do
    def change_order_type(self, ref_key: str):
        # if not CMS.user_access(self.db_naryad, 'аутсорс_смена_типа_заявки', F.user_full_namre()): #todo вкл
        #     return
        tbl = self.ui.tbl_list_zayv
        current_order = self.tree_objs2.first_object()
        if not current_order:
            return
        cur_order_row = current_order.current
        # cur_order_row = CQT.get_dict_line_form_tbl(tbl)
        old_doc_type = cur_order_row.get('НаименованиеДокумента')
        new_doc_type = self.get_doc_type_by_order_info()


        if new_doc_type is None:
            return CQT.msgbox('Не удалось вычислить новый тип документа')
        if old_doc_type is None:
            return CQT.msgbox('Не удалось вычислить старый тип документа')
        if old_doc_type == new_doc_type:
            return CQT.msgbox('Тип документа не изменился')
        code, response = self.erp_client.get_document(old_doc_type, ref_key=ref_key, fields=['DeletionMark'])
        if code != 200:
            return
        if not isinstance(response, dict) or 'DeletionMark' not in response:
            return

        if not response['DeletionMark']:
            if not CQT.msgboxgYN(f'Докусент с типом {old_doc_type!r} будет поставлен на удаление и создан документ {new_doc_type!r} продолжить?'):
                return
            if not self.mark_remove_order_in_ERP(remove_order_mes=False):
                return CQT.msgbox('Ошибка при попытке удалить документ ERP')
        self.create_an_order_in_ERP()

    def btn_create_or_change_order_type(self, *args):
        ref_key = self.get_order_ref_key()
        btn_state = self.ui.btn_add_line.text()
        if btn_state == BtnAddLineState.CHANGE_TYPE.value:
            self.change_order_type(ref_key)
        if btn_state == BtnAddLineState.EDIT_DOC.value:
            self.edit_an_order_in_ERP()
        if btn_state == BtnAddLineState.CREATE_DOC.value:
            self.create_an_order_in_ERP()


    def get_current_tree(self, root: ObjBranch):
        data = root.child.get_data()
        struct = root.current
        for item in data:
            child = root.child
            mutable_item = self.inject_params_into_object_item(child, item)
            child.current = mutable_item
            struct.setdefault(f'_{child.name}', list()).append(self.get_current_tree(child))
        return struct

    @CQT.onerror
    def edit_an_order_in_ERP(self, *args):
        tbl = self.ui.tbl_list_zayv
        current_row = tbl.currentRow()
        order_object = self.tree_objs2.first_object()
        cur_order = order_object.current
        if not cur_order:
            CQT.msgbox('Сначала необходимо выбрать заявку')
            return
        order_id = cur_order['Номер_заявки']
        order = CSQ.custom_request_c(
            CFG.Config.project.db_naryad,
            f'SELECT Ref_Key, НаименованиеДокумента FROM outsource WHERE id = {order_id}',
            rez_dict=True,
            one=True
        )
        ref_key = order['Ref_Key']
        result = self.get_current_tree(self.tree_objs2.first_object())
        nomen_dse = [dse for naryad in result['_Наряд'] for dse in naryad['_ДСЕ']]
        key_use_mats = 'Использование сырья'
        a = self.DICT_EMPLOEE_FULL.get(F.user_full_namre())
        dep_uuid = self.erp_client.department_by_individual_id(a.get('ID_ФизЛица'))
        if len(dep_uuid) >= 1:
            dep_uuid = dep_uuid[0]
        else:
            return CQT.msgbox('Не удалось опрелить подразделение')
        data_for_order = {
            'Комментарий': cur_order.get('Комментарий'),
            'Менеджер_Key': dep_uuid.get('Ref_Key'),
            'Автор_Key': dep_uuid.get('Ref_Key'),
            'Подразделение_Key': dep_uuid.get('Подразделение_Key'),
            'ЖелаемаяДатаПоступления': cur_order.get('ЖелаемаяДатаПоступления'),
            'ДополнительнаяИнформация': cur_order.get('ДополнительнаяИнформация')
        }
        code = 500
        doc_name = ''
        response = None
        match cur_order.get(key_use_mats):
            case 'Давальческое':
                doc_name = 'Document_ЗаказПереработчику2_5'
                docs = self.get_docs(order_id, doc_name)
                code, response = self.erp_client.edit_order(docs=docs, doc_name=doc_name, data_for_order=data_for_order, nomen_dse=nomen_dse, ref_key=ref_key)
            case 'Поставщика':
                doc_name = 'Document_ЗаказПоставщику'
                docs = self.get_docs(order_id, doc_name)
                code, response = self.erp_client.edit_order(docs=docs, doc_name=doc_name, data_for_order=data_for_order, nomen_dse=nomen_dse, ref_key=ref_key)
            case _:
                CQT.msgbox(f'Поле {key_use_mats!r} не заполнено')
        if code == 201:
            ref_key_new_order = response['Ref_Key']
            code_new_order = response['Number']
            date_new_order = response['Date']
            self.erp_client.mark_author(ref_key_new_order, doc_name, data_for_order.get('Автор'))
            order_id = cur_order.get('Номер_заявки')
            response = CSQ.custom_request_c(
                CFG.Config.project.db_naryad,
                f'UPDATE outsource SET Ref_Key = {ref_key_new_order!r}, НаименованиеДокумента = {doc_name!r} WHERE id = {order_id}'
            )
            if response:
                domain = 'srv-1c:3541'
                base_name = self.erp_client.client.srv_name
                nomen_endpoint = '.'.join(doc_name.split('_', 1))
                link = OERP.make_reference_on_erp_entity(
                    domain=domain,
                    base_name=base_name,
                    nomen_endpoint=nomen_endpoint,
                    ref_key_odata=ref_key_new_order
                )
                return CQT.msgboxg_get_table(
                    self,
                    'Документ успешно создан',
                    dict_or_list=[
                        {'Поле': 'Ссылка', 'Значение': f'{link}|Открыть документ в erp'},
                        {'Поле': 'Код', 'Значение': code_new_order},
                        {'Поле': 'Дата', 'Значение': date_new_order},
                    ],
                    load_links=True,
                    styleSheet=CQT.ERP_CSS,
                    show_filtr=False
                )
            CQT.msgbox('Ошибка. Не удалось создать заявку')

    @CQT.onerror
    def create_an_order_in_ERP(self, *args):
        def g_data(cur_dic, obj: ObjBranch):
            cur_dic = self.inject_params_into_object_item(obj, cur_dic)
            obj.current = cur_dic
            lst_items = obj.child.get_data()
            for i in lst_items:
                # params = self.get_object_params(obj.child.name)
                # i.update({field.get('Поле'): field.get('value') for field in params})
                g_data(i, obj.child)
            cur_dic['_' + obj.child.name] = lst_items
        tbl = self.ui.tbl_list_zayv
        current_row = tbl.currentRow()
        order_object = self.tree_objs2.first_object()
        cur_order = order_object.current
        if not cur_order:
            CQT.msgbox('Сначала необходимо выбрать заявку')
            return
        # cur_order = CQT.get_dict_line_form_tbl(self.ui.tbl_list_zayv)
        order_id = cur_order['Номер_заявки']
        # tree = TreeObjs()
        # order = tree.first_object()
        # order.current = cur_order
        # data = []
        # items = order.child.get_data()
        # result = {}
        result = self.get_current_tree(self.tree_objs2.first_object())
        # if '_Наряд' not in result:
        #     return
        # nar = order.child
        # for item in items:
        #     g_data(item, nar)
        #     data.append(item)
        nomen_dse = [dse for naryad in result['_Наряд'] for dse in naryad['_ДСЕ']]
        key_use_mats = 'Использование сырья'
        a = self.DICT_EMPLOEE_FULL.get(F.user_full_namre())
        dep_uuid = self.erp_client.department_by_individual_id(a.get('ID_ФизЛица'))
        if len(dep_uuid) >= 1:
            dep_uuid = dep_uuid[0]
        else:
            return CQT.msgbox('Не удалось опрелить подразделение')
        data_for_order = {
            'Комментарий': cur_order.get('Комментарий'),
            'Менеджер_Key': dep_uuid.get('Ref_Key'),
            'Автор_Key': dep_uuid.get('Ref_Key'),
            'Подразделение_Key': dep_uuid.get('Подразделение_Key'),
            'ЖелаемаяДатаПоступления': cur_order.get('ЖелаемаяДатаПоступления'),
            'ДополнительнаяИнформация': cur_order.get('ДополнительнаяИнформация')
        }
        code = 500
        doc_name = ''
        response = None
        match cur_order.get(key_use_mats):
            case 'Давальческое':
                doc_name = 'Document_ЗаказПереработчику2_5'
                docs = self.get_docs(order_id, doc_name)
                code, response = self.erp_client.create_order(docs=docs, doc_name=doc_name, data_for_order=data_for_order, nomen_dse=nomen_dse)
            case 'Поставщика':
                doc_name = 'Document_ЗаказПоставщику'
                docs = self.get_docs(order_id, doc_name)
                code, response = self.erp_client.create_order(docs=docs, doc_name=doc_name, data_for_order=data_for_order, nomen_dse=nomen_dse)
            case _:
                CQT.msgbox(f'Поле {key_use_mats!r} не заполнено')
        if code == 201:
            ref_key_new_order = response['Ref_Key']
            code_new_order = response['Number']
            date_new_order = response['Date']
            self.erp_client.mark_author(ref_key_new_order, doc_name, data_for_order.get('Автор'))
            order_id = cur_order.get('Номер_заявки')
            response = CSQ.custom_request_c(
                CFG.Config.project.db_naryad,
                f'UPDATE outsource SET Ref_Key = {ref_key_new_order!r}, НаименованиеДокумента = {doc_name!r} WHERE id = {order_id}'
            )
            if response:
                domain = 'srv-1c:3541'
                base_name = self.erp_client.client.srv_name
                nomen_endpoint = '.'.join(doc_name.split('_', 1))
                link = OERP.make_reference_on_erp_entity(
                    domain=domain,
                    base_name=base_name,
                    nomen_endpoint=nomen_endpoint,
                    ref_key_odata=ref_key_new_order
                )
                return CQT.msgboxg_get_table(
                    self,
                    'Документ успешно создан',
                    dict_or_list=[
                        {'Поле': 'Ссылка', 'Значение': f'{link}|Открыть документ в erp'},
                        {'Поле': 'Код', 'Значение': code_new_order},
                        {'Поле': 'Дата', 'Значение': date_new_order},
                    ],
                    load_links=True,
                    styleSheet=CQT.ERP_CSS,
                    show_filtr=False
                )
            CQT.msgbox('Ошибка. Не удалось создать заявку')

    @CQT.onerror
    @refresh_main_tbl_after_do
    def mark_remove_order_in_ERP(self, *args, remove_order_mes: bool = True) -> bool:
        # cur_order = CQT.get_dict_line_form_tbl(self.ui.tbl_list_zayv)
        order_object = self.tree_objs2.first_object()
        cur_order = order_object.current
        pk_order = cur_order.get('Номер_заявки')
        state = bool(int(cur_order.get('НаУдаление', 0)))
        new_state = not state
        order = CSQ.custom_request_c(
            CFG.Config.project.db_naryad,
            f'SELECT Ref_Key, НаименованиеДокумента FROM outsource WHERE id = {pk_order}',
            rez_dict=True,
            one=True
        )
        if isinstance(order, dict):
            doc_name = order.get('НаименованиеДокумента')
            if isinstance(order.get('Ref_Key'), str) and len(order['Ref_Key']) == 36 and doc_name:
                code, body = self.erp_client.mark_remove_order(order['Ref_Key'], doc_name, mark=new_state)
                ref_key = order['Ref_Key']

                if code != 200 or body.get('DeletionMark') != new_state:
                    return CQT.msgbox('Произошла ошибка при попытке удалить документ в ERP')
                change_logger('Поставил на удаление документ ERP', f'Документ: {doc_name} Ref_Key: {ref_key}')
            if not remove_order_mes:
                return True
            else:
                response = CSQ.custom_request_c(self.db_naryad, f'UPDATE outsource SET НаУдаление = {int(new_state)} WHERE id = {pk_order}')
                if response:
                    if not new_state:
                        change_logger('Снял метку НаУдаление в заявке на аутсорс МЕС', f'ID: {pk_order}')
                        CQT.msgbox('Заявка успешно снята с удаления')
                        return True
                    change_logger('Пометил на удаление заявку на аутсорс МЕС', f'ID: {pk_order}')
                    CQT.msgbox('Заявка успешно удалена!')
                    return True
            return CQT.msgbox('Ошибка при попытке удалить заявку в МЕС')

    def decor_table(self):
        match self.current_obj_name:
            case 'Заявка': self.decor_app()

    def decor_app(self):
        tbl = self.ui.tbl_list_zayv
        # dict_by_status = aggregate_by_key(self.DICT_PARAMS_BY_OBJECT.get('Заявка', []), 'Статус')
        # for idx, row in enumerate(result):
        #     cur_status = self.order_status[1]
        #     for idx, status in enumerate(self.order_status):
        #         if cur_status_params := dict_by_status.get(status, []):
        #             if all(row.get(field.get('Поле')) for field in cur_status_params):
        #                 cur_status = self.order_status[idx + 1 if idx < len(self.order_status) else len(self.order_status) - 1]
        #     row['Статус'] = cur_status
        nk_create = CQT.num_col_by_name_c(tbl, 'Создать')
        nk_docs = CQT.num_col_by_name_c(tbl, 'Docs')
        nk_status = CQT.num_col_by_name_c(tbl, 'Статус')
        # for row in range(tbl.rowCount()):
        #     # cur_status = tbl.item(row, nk_status).text()
        # #     # TODO создание документа erp
        # #     # TODO просмотр прикрепленных документов docs
        #     # if cur_status and self.order_status.index(cur_status) >= 2:
        #     # CQT.add_btn(tbl, row, nk_create, 'ERP', conn_func_checked_row_col=self.create_an_order_in_ERP)
        #     CQT.add_btn(tbl, row, nk_docs, 'DOC', conn_func_checked_row_col=self.get_docs)

    def fill_new_zayav(self):
        tbl_zayav = self.ui.tbl_new_order
        tbl_zayav.setRowCount(0) or tbl_zayav.setRowCount(1)

        params = self.DICT_PARAMS_BY_OBJECT.get('Заявка')
        data = {}
        for item in params:
            text = item.get('Поле')
            if item.get('Статус') == 'Создание':
                data[text] = ''
        CQT.fill_wtabl([data], tbl_zayav, min_width_col=120)
        self.set_color_departments(tbl_zayav)
        tbl_nar = self.ui.tbl_list_naryads

        data = get_wo_autsorce()
        CQT.fill_wtabl(data, tbl_nar)
        CQT.fill_filtr_c(self, self.ui.tbl_list_naryads_filtr, tbl_nar, hidden_scroll=True)

    def click_new_app(self):
        self.ui.tabWidget.setCurrentIndex(1)
        self.fill_new_zayav()
        tbl = self.ui.tbl_list_naryads
        tbl_filter = self.ui.tbl_list_naryads_filtr
        naryads = get_wo_autsorce()
        CQT.fill_wtabl(naryads, tbl)
        CMS.fill_filtr_c(self, tbl_filter, tbl, hidden_scroll=True)

    def choose_naryad(self, *args, **kwargs):
        main_tbl = self.ui.tbl_list_naryads
        chse_tbl = self.ui.tbl_list_chosed_naryads
        row_val = CQT.get_dict_line_form_tbl(main_tbl)
        chse_val = CQT.list_from_wtabl_c(chse_tbl, hat_c=False, rez_dict=True)
        if not row_val:
            return
        params = self.DICT_PARAMS_BY_OBJECT.get('Наряд')
        added_params = {'Номер_наряда': row_val.get('Пномер')}

        for param in params:
            key = param.get('Поле')
            added_params[key] = ''
        # added_params.update(row_val)
        CQT.fill_wtabl([added_params, *chse_val], self.ui.tbl_list_chosed_naryads, min_width_col=120)
        CQT.fill_filtr_c(self, self.ui.tbl_list_chosed_naryads_filtr, self.ui.tbl_list_chosed_naryads, hidden_scroll=True)
        chse_tbl.selectRow(0)
        self.on_select_choosed_naryad()

    def get_param_line_by_field(self, obj_name: str, field: str):
        field_lines = self.DICT_PARAMS_BY_OBJECT.get(obj_name, [])
        for val in field_lines:
            if val.get('Поле') == field:
                return val

    def apply_edit(self, *args):
        if self.info_focus_table == 'param':
            self.apply_edit_param(*args)
        if self.info_focus_table == 'order':
            self.apply_edit_order(*args)
        if self.info_focus_table == 'naryad':
            self.apply_edit_naryad(*args)
        if self.info_focus_table == 'main':
            self.apply_edit_main()

    def apply_edit_main(self, *args):
        tbl = self.ui.tbl_list_zayv
        current_row = tbl.currentRow()
        tbl_val = dict_from_info_table(self.ui.tbl_infozayav)
        old_tbl_state = CQT.list_from_wtabl_c(tbl, hat_c=False, rez_dict=True)
        keys = []
        pks = []
        for obj in self.obj_keys[:self.obj_current_row + 1]:
            inst: ObjBranch = self.tree_objs2[obj]
            keys.append(inst.param_id)
            pks.append(inst.current_id)
        list_values = []
        changes = []
        for key, val in tbl_val.items():
            param_info = self.info_params.get(key)
            # param_info = self.get_param_line_by_field(self.current_obj_name, key)
            if param_info is not None and self.department in param_info.department: # todo правка параметров по ответствтенному за поле подр
                old_value = old_tbl_state[self.current_main_row][key]
                if old_value != val:
                    list_values.append([param_info.pk, val, *pks])
                    changes.append(f'<{key}> Было: {old_value} Стало: {val}')
                    old_tbl_state[self.current_main_row][key] = val
        if list_values:
            que = ', '.join('?' for _ in list_values[0])
            query = f"""
                INSERT OR REPLACE INTO out_params (param_id, value, {', '.join(keys)})
                VALUES ({que});
            """
            if CSQ.custom_request_c(self.db_naryad, query, list_of_lists_c=list_values):
                change_logger(action='Изменение данных', data='\n'.join(changes))
                self.set_objs(current_row=self.obj_current_row)
                tbl.selectRow(current_row)
                CQT.msgbox('Изменения в "%s" успешно сохранены' % self.current_obj_name)

    def apply_edit_order(self, *args):
        tbl = self.ui.tbl_new_order
        tbl_val = dict_from_info_table(self.ui.tbl_infozayav)
        old_val = CQT.get_dict_line_form_tbl(tbl, 0)
        for key, val in tbl_val.items():
            old_val[key] = val
        CQT.fill_wtabl([old_val], tbl)

    def apply_edit_naryad(self, *args):
        tbl = self.ui.tbl_infozayav
        tbl_val = dict_from_info_table(tbl)
        tbl_naryad = self.ui.tbl_list_chosed_naryads
        tbl_naryad_val = CQT.list_from_wtabl_c(tbl_naryad, hat_c=False, rez_dict=True)
        for key, val in tbl_val.items():
            tbl_naryad_val[self.current_info_naryad][key] = val
        CQT.fill_wtabl(tbl_naryad_val, tbl_naryad)
        CQT.fill_filtr_c(self, self.ui.tbl_list_chosed_naryads_filtr, self.ui.tbl_list_chosed_naryads, hidden_scroll=True)

    def apply_edit_param(self, *args):
        # if not CMS.user_access(self.db_naryad, 'аутсорс_создание_корректировка_дополнительных_полей', F.user_full_namre()): # todo создать права
        #     return
        tbl = self.ui.tbl_infozayav
        CQT.get_dict_line_form_tbl(tbl)
        tbl_val = dict_from_info_table(tbl)
        department = tbl_val.get('Подразделение')
        field = tbl_val.get('Поле')
        note = tbl_val.get('Описание')
        status = tbl_val.get('Статус')
        typ_field = tbl_val.get('Тип')
        default_values = tbl_val.get('По_умолчанию')
        obj_name = tbl_val.get('Объект')
        if any(len(value) < 4 for value in (field, note)):
            return CQT.msgbox('Строки "Поле" и "Описание" должны иметь более 4 символов')
        if pk := tbl_val.get('Пномер'):
            query = f'''
                UPDATE out_params_info 
                SET Подразделение = ?, Поле = ?, Описание = ?, Статус = ?, Тип = ?, По_умолчанию = ?, Объект = ?
                WHERE Пномер = {pk}
            '''
            action = 'обновлен'
        else:
            action = 'создан'
            query = '''
                INSERT INTO out_params_info(Подразделение, Поле, Описание, Статус, Тип, По_умолчанию, Объект) 
                VALUES (?, ?, ?, ?, ?, ?, ?)
            '''
        resp = CSQ.custom_request_c(self.db_naryad, query, list_of_lists_c=[[
            department, field, note, status, typ_field, default_values, obj_name]])
        if resp:
            self.fill_params_table()
            CQT.msgbox(f'Параметр успешно {action}')


    def clear_selected_naryads(self, *args):
        self.ui.tbl_list_chosed_naryads.setRowCount(0)
        self.ui.tbl_list_chosed_naryads.setColumnCount(0)

    def dump_excel(self, *args):
        tbl = self.ui.tbl_list_zayv
        if tbl.currentRow() == -1:
            CQT.msgbox('Сначала выделите нужную заявку')
            return
        fn, _ = QtWidgets.QFileDialog.getSaveFileName(self, '', 'zxc', '(*.xlsx)')
        header = [
            'Тип заявки', 'Заказ на производство', 'Материал/Наименование',
            'Ед. изм.', 'Кол-во', 'Материал ( если не указано в чертеже )',
            'Материал норма, кг. на кол-во', 'Операция', 'Требуется ли предоставление ответной детали?',
            'нормы времени по операции на количество, Мин', 'Вид заготвки ( сортамент полуфабрикат )', 'Маркировка деталей',
            'Примечание', 'Виды упаковки', 'Виды тары',
            'Необходимая дата для ответа ( проработка 3-5 дней )', 'Ннеобходимая дата поставки на склад Пауэрз', 'Дата готовности заготовки к передаче',
            'Использование сырья', 'Требуемые документы от поставщика'
        ]
        wb = Workbook(write_only=False, iso_dates=False)
        ws = wb.active
        ws.title = 'abc'

        order_vals = CQT.get_dict_line_form_tbl(tbl)
        if not isinstance(order_vals, dict): return
        naryad_lst = select_naryads(Номер_заявки=order_vals.get('Номер_заявки'))
        ws.append(header)

        orange = PatternFill(start_color="ed7d31", end_color="ed7d31", fill_type="solid")
        green = PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")

        for idx, cell in enumerate(ws[1]):
            color = orange
            if idx in {0, 1, 7, 9, 12, 15}:
                color = green
            cell.fill = color

            cell.fill = color
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for naryad in naryad_lst:
            dse_lst = naryad.get('ДСЕ', '').split('|')
            oper_lst = naryad.get('Операции', '').split('|')
            for dse, oper in zip(dse_lst, oper_lst):
                name, nn = dse.split('$')
                oper_code, oper_name = oper.split('$')
                tmp = [
                    '-', '-', f'{nn} {name}',
                    'шт', '-', '-',
                    '-', oper_name, order_vals.get('Ответная_деталь', '-'),
                    '-', '-', order_vals.get('Маркировка_деталей', '-'),
                    '-', naryad.get('Упаковка', '-'), naryad.get('Тара', '-'),
                    order_vals.get('Дата ответа', '-'), naryad.get('Дата поставки на склад', '-'), naryad.get('Дата передачи', '-'),
                    order_vals.get('Использование сырья', '-'), order_vals.get('Требуемые документы поставщика', '-')
                ]
                ws.append(tmp)

        for row in ws:
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    cell.border = thin_border
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(fn)
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = True
        excel.Workbooks.Open(fn)


if __name__ == "__main__":
    app = QtWidgets.QApplication(['', '--no-sandbox'])
    application = MainWindow()
    application.showMaximized()
    sys.exit(app.exec())
