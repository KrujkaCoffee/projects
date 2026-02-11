from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _safe_filename(name: str, repl: str = "_") -> str:
    bad = r'<>:"/\\|?*'
    out = []
    for ch in (name or "").strip():
        if ch in bad or ord(ch) < 32:
            out.append(repl)
        else:
            out.append(ch)
    s = "".join(out).strip()
    return s if s else "report"


def _is_empty(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, str):
        return val.strip() == ""
    try:
        # NaN / Inf
        import math
        if isinstance(val, (int, float)) and (math.isnan(val) or math.isinf(val)):
            return True
    except Exception:
        pass
    return False


def _try_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(",", ".")
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def _as_list(v):
    if v is None:
        return None
    if isinstance(v, str):
        return [x.strip() for x in v.split(",") if x.strip()]
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    return None


@dataclass(frozen=True)
class TechReportCfg:
    """
    Конфиг отчёта.
    Все поля опциональные; если чего-то нет — используются безопасные дефолты.
    """
    # key -> bool (если False - не включать)
    fields: Dict[str, bool] | None = None
    # group_id/group_name -> bool (если False - не включать целиком)
    groups: Dict[str, bool] | None = None

    transpose_enabled: bool = True
    # теги серий для транспонирования: n, out, k, ...
    transpose_tags: List[str] | None = None
    # дополнительные фильтры транспонирования (опционально):
    # - transpose_groups: группа -> bool (если False - серии в группе НЕ транспонировать)
    # - transpose_bases: allow-list баз (если None - транспонировать все базы; если [] - ни одну)
    transpose_groups: Dict[str, bool] | None = None
    transpose_bases: List[str] | None = None
    # регулярка для распознавания серий: <base>_<tag><idx>
    suffix_regex: str = r"^(?P<base>.+)_(?P<tag>[A-Za-z]+)(?P<idx>\d+)$"

    transpose_tag_bases: List[str] | None = None
    transpose_num_prefixes: List[str] | None = None
    transpose_numfix: List[str] | None = None


def cfg_from_any(raw: Any) -> TechReportCfg:
    """
    Преобразует любой словарь (например из cust_data.tech_report_cfg) в TechReportCfg
    без жёсткой зависимости от структуры UI.
    """
    if not isinstance(raw, dict):
        return TechReportCfg()

    # разные возможные формы
    fields = raw.get("fields") or raw.get("field_visibility") or raw.get("visible_fields")
    if isinstance(fields, dict):
        fields = {str(k): bool(v) for k, v in fields.items()}
    else:
        fields = None

    groups = raw.get("groups") or raw.get("group_visibility") or raw.get("visible_groups")
    if isinstance(groups, dict):
        groups = {str(k): bool(v) for k, v in groups.items()}
    else:
        groups = None

    transpose_enabled = raw.get("transpose_enabled")
    if transpose_enabled is None:
        transpose_enabled = raw.get("transpose", {}).get("enabled") if isinstance(raw.get("transpose"), dict) else True
    transpose_enabled = bool(transpose_enabled)

    tags = raw.get("transpose_tags")
    if tags is None and isinstance(raw.get("transpose"), dict):
        tags = raw["transpose"].get("tags")
    if isinstance(tags, str):
        # "n,out"
        tags = [t.strip() for t in tags.split(",") if t.strip()]
    if isinstance(tags, list):
        tags = [str(t).strip() for t in tags if str(t).strip()]
    else:
        tags = None

    suffix_regex = raw.get("suffix_regex")
    if suffix_regex is None and isinstance(raw.get("transpose"), dict):
        suffix_regex = raw["transpose"].get("suffix_regex")
    if not isinstance(suffix_regex, str) or not suffix_regex.strip():
        suffix_regex = TechReportCfg.suffix_regex

    # ---- optional transpose filters ----
    t_groups = raw.get("transpose_groups")
    if t_groups is None and isinstance(raw.get("transpose"), dict):
        t_groups = raw["transpose"].get("groups")
    if isinstance(t_groups, dict):
        t_groups = {str(k): bool(v) for k, v in t_groups.items()}
    else:
        t_groups = None

    t_bases = raw.get("transpose_bases")
    if t_bases is None and isinstance(raw.get("transpose"), dict):
        t_bases = raw["transpose"].get("bases")
    # поддержим формы: list[str] или dict[str,bool]
    if isinstance(t_bases, dict):
        t_bases = [str(k) for k, v in t_bases.items() if bool(v)]
    if isinstance(t_bases, str):
        # "a,b,c"
        t_bases = [x.strip() for x in t_bases.split(",") if x.strip()]
    if isinstance(t_bases, list):
        t_bases = [str(x).strip() for x in t_bases if str(x).strip()]
    else:
        t_bases = None

    tag_bases = _as_list(raw.get("transpose_tag_bases"))
    num_prefixes = _as_list(raw.get("transpose_num_prefixes"))
    numfix = _as_list(raw.get("transpose_numfix"))

    return TechReportCfg(
        fields=fields,
        groups=groups,
        transpose_enabled=transpose_enabled,
        transpose_tags=tags,
        transpose_groups=t_groups,
        transpose_bases=t_bases,
        suffix_regex=suffix_regex,
        transpose_tag_bases=tag_bases,
        transpose_num_prefixes=num_prefixes,
        transpose_numfix=numfix,
    )


# ------------------------- main build -------------------------

def build_tech_report_xlsx(
    *,
    report_name: str,
    input_rows: List[dict],
    calculated: Dict[str, Any],
    output_params: Dict[str, dict],
    errors: Optional[List[dict]] = None,
    save_dir: str,
    module_alias: str = "silencer",
    cfg_raw: Any = None,
) -> str | bool:
    """
    Создаёт xlsx и возвращает путь или False.

    input_rows — результат common_funcs.datatable_to_dicts(input_table)
    calculated — dict всех рассчитанных значений (prepare_calc_new_data())
    output_params — метаописание параметров (header, dimension, group_name, accuracy, comment, ...)
    """

    cfg = cfg_from_any(cfg_raw)

    if not os.path.isdir(save_dir):
        os.makedirs(save_dir, exist_ok=True)

    safe = _safe_filename(report_name)
    file_name = f"{safe}_tech.xlsx"
    path = os.path.join(save_dir, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Технологический отчёт"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    group_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_cell(r, c, value, *, font=None, fill=None, align=None, border_=None):
        cell = ws.cell(row=r, column=c, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if border_:
            cell.border = border_
        return cell

    row = 1
    set_cell(row, 1, "Технологический отчёт", font=title_font)
    row += 1
    set_cell(row, 1, report_name, font=bold)
    row += 2

    set_cell(row, 1, "Входные данные", font=bold)
    row += 1

    in_headers = ["Параметр", "Значение", "Ед.изм"]
    for col, h in enumerate(in_headers, start=1):
        set_cell(row, col, h, font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
    row += 1

    for it in input_rows or []:

        param = it.get("Параметр", "")
        val = it.get("Значение", "")
        dim = it.get("Ед.изм", "")
        if val in in_headers:
            continue
        set_cell(row, 1, param, border_=border)
        set_cell(row, 2, val, border_=border)
        set_cell(row, 3, dim, border_=border)
        row += 1

    row += 2

    set_cell(row, 1, "Результаты расчёта", font=bold)
    row += 1

    out_base_headers = ["Параметр", "Значение", "Ед.изм", "Комментарий"]
    for col, h in enumerate(out_base_headers, start=1):
        set_cell(row, col, h, font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
    row += 1

    grouped: Dict[str, List[str]] = {}

    def _default_visible(meta: dict) -> bool:
        """Определяем базовую видимость из output_params[key]['view'].

        Поддерживаем форматы:
        - view: bool
        - view: {"visible": bool, ...}
        - отсутствует -> True
        """
        view = meta.get("view", True)
        if isinstance(view, bool):
            return view
        if isinstance(view, dict):
            return bool(view.get("visible", True))
        return True

    for key, val in (calculated or {}).items():
        if key not in output_params:
            continue

        meta = output_params[key] or {}
        group_name = str(meta.get("group_name", "") or "").strip()

        if cfg.groups is not None:
            if group_name in cfg.groups and not cfg.groups[group_name]:
                continue

        visible = _default_visible(meta)
        if cfg.fields is not None and key in cfg.fields:
            visible = bool(cfg.fields[key])
        if not visible:
            continue

        if _is_empty(val):
            continue

        grouped.setdefault(group_name, []).append(key)

    group_names = sorted(grouped.keys(), key=lambda s: s.lower())

    suffix_re = re.compile(cfg.suffix_regex)
    allowed_tags = set(t.lower() for t in (cfg.transpose_tags or ["n"]))  # default n
    allowed_bases = None
    if cfg.transpose_bases is not None:
        allowed_bases = set(str(b) for b in cfg.transpose_bases)

    for gname in group_names:
        keys = grouped[gname]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        set_cell(row, 1, gname if gname else "Без группы", font=bold, fill=group_fill, border_=border)
        row += 1

        def _sort_key(k: str):
            meta = output_params.get(k, {}) or {}
            return str(meta.get("header", k)).lower()

        keys_sorted = sorted(keys, key=_sort_key)

        series: Dict[str, Dict[str, Dict[int, str]]] = {}  # tag -> base -> idx -> key
        normal: List[str] = []

        for k in keys_sorted:
            m = suffix_re.match(k)
            if cfg.transpose_enabled and m:
                if cfg.transpose_groups is not None:
                    if gname in cfg.transpose_groups and not cfg.transpose_groups[gname]:
                        normal.append(k)
                        continue
                tag = m.group("tag").lower()
                if (not allowed_tags) or (tag in allowed_tags):
                    base = m.group("base")
                    if allowed_bases is not None and base not in allowed_bases:
                        normal.append(k)
                        continue
                    idx = int(m.group("idx"))
                    series.setdefault(tag, {}).setdefault(base, {})[idx] = k
                    continue
            normal.append(k)

        for k in normal:
            meta = output_params.get(k, {}) or {}
            header = meta.get("header", k)
            dim = meta.get("dimension", "")
            comment = meta.get("comment") or ""

            v = calculated.get(k)
            f = _try_float(v)
            if f is not None and "accuracy" in meta:
                try:
                    f = round(f, int(meta.get("accuracy") or 0))
                except Exception:
                    pass
                v_out = f
            else:
                v_out = v

            set_cell(row, 1, header, border_=border)
            set_cell(row, 2, v_out, border_=border)
            set_cell(row, 3, dim, border_=border)
            set_cell(row, 4, comment, border_=border)
            row += 1

        for tag, base_map in series.items():
            if not base_map:
                continue

            if normal:
                row += 1

            # determine indices union
            idxs = sorted({idx for bm in base_map.values() for idx in bm.keys()})
            # header for block
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            set_cell(row, 1, f"Серии: _{tag}{{N}}", font=bold, fill=header_fill)
            row += 1

            # block header row: Parameter | tag1 | tag2 | ... | Unit
            set_cell(row, 1, "Параметр", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            for j, idx in enumerate(idxs, start=2):
                set_cell(row, j, f"{tag}{idx}", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            set_cell(row, 2 + len(idxs), "Ед.изм", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            row += 1

            # write each base as a row
            for base, idx_map in sorted(base_map.items(), key=lambda x: x[0].lower()):
                # pick meta from the smallest idx key for header/dim
                first_key = idx_map[sorted(idx_map.keys())[0]]
                meta0 = output_params.get(first_key, {}) or {}
                header = meta0.get("header", base)
                dim = meta0.get("dimension", "")

                set_cell(row, 1, header, border_=border)
                for j, idx in enumerate(idxs, start=2):
                    k = idx_map.get(idx)
                    if not k:
                        set_cell(row, j, "", border_=border)
                        continue
                    v = calculated.get(k)
                    f = _try_float(v)
                    if f is not None and "accuracy" in meta0:
                        try:
                            f = round(f, int(meta0.get("accuracy") or 0))
                        except Exception:
                            pass
                        set_cell(row, j, f, border_=border)
                    else:
                        set_cell(row, j, v, border_=border)

                set_cell(row, 2 + len(idxs), dim, border_=border)
                row += 1

            row += 1  # gap after block

        row += 1  # gap after group

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 60)

    ws.freeze_panes = "A5"

    if errors:
        ws_err = wb.create_sheet("Ошибки")
        ws_err["A1"] = "Параметры с ошибками расчёта"
        ws_err["A1"].font = title_font
        ws_err.append(["Параметр", "Сообщение"])
        ws_err["A2"].font = bold
        ws_err["B2"].font = bold
        for err in errors:
            hdr = err.get("header") or err.get("name") or ""
            msg = err.get("Exception") or err.get("msg") or err.get("error") or ""
            ws_err.append([str(hdr), str(msg)])
        ws_err.column_dimensions["A"].width = 40
        ws_err.column_dimensions["B"].width = 80

    wb.save(path)
    return path
