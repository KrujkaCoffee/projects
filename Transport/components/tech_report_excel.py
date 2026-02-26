from __future__ import annotations

import os
import re
import math
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class TechReportBuilder:
    def clean_forbidden_chars(self, name: str, repl: str = "_") -> str:
        bad = r'<>:"/\\|?*'
        out = []
        for ch in (name or "").strip():
            if ch in bad or ord(ch) < 32:
                out.append(repl)
            else:
                out.append(ch)
        s = "".join(out).strip()
        return s if s else "report"


def _is_empty(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, str):
        return val.strip() == ""
    try:
        if isinstance(val, (int, float)) and (math.isnan(val) or math.isinf(val)):
            return True
    except Exception:
        pass
    return False


def _try_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(",", ".")
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def _as_list(v: Any) -> Optional[List[str]]:
    if v is None:
        return None
    if isinstance(v, str):
        return [x.strip() for x in v.split(",") if x.strip()]
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    return None


def _slug_base(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^0-9a-zа-я]+", "_", s, flags=re.IGNORECASE)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "series"


_FREQ_RE = re.compile(
    r"(?i)(?:гц|hz)\s*(?P<a>\d+(?:[.,]\d+)?)|(?P<b>\d+(?:[.,]\d+)?)\s*(?:гц|hz)"
)


def _extract_freq(header: str, comment: str) -> Optional[float]:
    """Ищем частоту (31,5 / 250 / 1000) в header или comment."""
    for s in (header or "", comment or ""):
        m = _FREQ_RE.search(str(s))
        if not m:
            continue
        raw = m.group("a") or m.group("b")
        if not raw:
            continue
        try:
            return float(raw.replace(",", "."))
        except Exception:
            return None
    return None


def _strip_freq_suffix(header: str) -> str:
    """Убираем хвост вида ', Гц 250' или '250 Гц' если он стоит в конце header."""
    s = (header or "").strip()
    if not s:
        return s
    s2 = re.sub(r"(?:,\s*)?(?:гц|hz)\s*\d+(?:[.,]\d+)?\s*$", "", s, flags=re.IGNORECASE).strip()
    if s2 != s:
        return s2.strip(" ,")
    s2 = re.sub(r"\d+(?:[.,]\d+)?\s*(?:гц|hz)\s*$", "", s, flags=re.IGNORECASE).strip()
    return s2.strip(" ,")


@dataclass(frozen=True)
class TechReportCfg:
    """Конфиг тех-отчёта."""
    fields: Optional[Dict[str, bool]] = None
    groups: Optional[Dict[str, bool]] = None

    transpose_enabled: bool = True
    transpose_mode: str = "auto"
    transpose_tags: Optional[List[str]] = None
    transpose_groups: Optional[Dict[str, bool]] = None
    transpose_bases: Optional[List[str]] = None
    suffix_regex: str = r"^(?P<base>.+)_(?P<tag>[A-Za-z]+)(?P<idx>\d+)$"

    debug_sheet: bool = False


def cfg_from_any(raw: Any) -> TechReportCfg:
    if not isinstance(raw, dict):
        return TechReportCfg()

    fields = raw.get("fields") or raw.get("field_visibility") or raw.get("visible_fields")
    fields = {str(k): bool(v) for k, v in fields.items()} if isinstance(fields, dict) else None

    groups = raw.get("groups") or raw.get("group_visibility") or raw.get("visible_groups")
    groups = {str(k): bool(v) for k, v in groups.items()} if isinstance(groups, dict) else None

    transpose_enabled = raw.get("transpose_enabled")
    if transpose_enabled is None and isinstance(raw.get("transpose"), dict):
        transpose_enabled = raw["transpose"].get("enabled")
    transpose_enabled = bool(True if transpose_enabled is None else transpose_enabled)

    mode = raw.get("transpose_mode")
    if mode is None and isinstance(raw.get("transpose"), dict):
        mode = raw["transpose"].get("mode")
    mode = str(mode or "auto").strip().lower()
    if mode == "header_n":
        mode = "freq"
    if mode not in ("auto", "suffix", "freq"):
        mode = "auto"

    tags = raw.get("transpose_tags")
    if tags is None and isinstance(raw.get("transpose"), dict):
        tags = raw["transpose"].get("tags")
    if isinstance(tags, str):
        tags = [t.strip() for t in tags.split(",") if t.strip()]
    if isinstance(tags, list):
        tags = [str(t).strip() for t in tags if str(t).strip()]
    else:
        tags = None

    suffix_regex = raw.get("suffix_regex")
    if suffix_regex is None and isinstance(raw.get("transpose"), dict):
        suffix_regex = raw["transpose"].get("suffix_regex")
    if not isinstance(suffix_regex, str) or not suffix_regex.strip():
        suffix_regex = TechReportCfg.suffix_regex

    t_groups = raw.get("transpose_groups")
    if t_groups is None and isinstance(raw.get("transpose"), dict):
        t_groups = raw["transpose"].get("groups")
    t_groups = {str(k): bool(v) for k, v in t_groups.items()} if isinstance(t_groups, dict) else None

    t_bases = raw.get("transpose_bases")
    if t_bases is None and isinstance(raw.get("transpose"), dict):
        t_bases = raw["transpose"].get("bases")
    if isinstance(t_bases, dict):
        t_bases = [str(k) for k, v in t_bases.items() if bool(v)]
    if isinstance(t_bases, str):
        t_bases = [x.strip() for x in t_bases.split(",") if x.strip()]
    if isinstance(t_bases, list):
        t_bases = [str(x).strip() for x in t_bases if str(x).strip()]
    else:
        t_bases = None

    debug_sheet = bool(raw.get("debug_sheet", False))

    return TechReportCfg(
        fields=fields,
        groups=groups,
        transpose_enabled=transpose_enabled,
        transpose_mode=mode,
        transpose_tags=tags,
        transpose_groups=t_groups,
        transpose_bases=t_bases,
        suffix_regex=suffix_regex,
        debug_sheet=debug_sheet,
    )


def build_tech_report_xlsx(
    *,
    report_name: str,
    input_rows: List[dict],
    calculated: Dict[str, Any],
    output_params: Dict[str, dict],
    errors: Optional[List[dict]] = None,
    save_dir: str,
    module_alias: str = "silencer",
    cfg_raw: Any = None,
) -> str | bool:
    """Создаёт xlsx и возвращает путь, либо False."""
    cfg = cfg_from_any(cfg_raw)

    if not os.path.isdir(save_dir):
        os.makedirs(save_dir, exist_ok=True)
    safe = TechReportBuilder().clean_forbidden_chars(report_name)
    file_name = f"{safe}_tech.xlsx"
    path = os.path.join(save_dir, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Технологический отчёт"

    debug_entries: list[dict[str, str]] = []

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    group_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_cell(r, c, value, *, font=None, fill=None, align=None, border_=None):
        cell = ws.cell(row=r, column=c, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if border_:
            cell.border = border_
        return cell

    def _default_visible(meta: dict) -> bool:
        view = meta.get("view", True)
        if isinstance(view, bool):
            return view
        if isinstance(view, dict):
            return bool(view.get("visible", True))
        return True

    def _push_debug(*, key: str, group: str, meta: dict, value_cell_addr: str):
        header = str(meta.get("header", key))
        comment = str(meta.get("comment") or "")
        debug_entries.append(
            {
                "key": str(key),
                "group": str(group),
                "header": header,
                "comment": comment,
                "addr": value_cell_addr,
            }
        )

    row = 1
    set_cell(row, 1, "Технологический отчёт", font=title_font)
    row += 1
    set_cell(row, 1, report_name, font=bold)
    row += 2

    set_cell(row, 1, "Входные данные", font=bold)
    row += 1
    in_headers = ["Параметр", "Значение", "Ед.изм"]
    for col, h in enumerate(in_headers, start=1):
        set_cell(row, col, h, font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
    row += 1
    for it in input_rows or []:
        param = it.get("Параметр", "")
        val = it.get("Значение", "")
        dim = it.get("Ед.изм", "")
        if not param: continue
        if val in in_headers:
            continue
        set_cell(row, 1, param, border_=border)
        set_cell(row, 2, val, border_=border)
        set_cell(row, 3, dim, border_=border)
        row += 1

    row += 2

    set_cell(row, 1, "Результаты расчёта", font=bold)
    row += 1
    out_base_headers = ["Параметр", "Значение", "Ед.изм", "Комментарий"]
    for col, h in enumerate(out_base_headers, start=1):
        set_cell(row, col, h, font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
    row += 1

    grouped: dict[str, list[str]] = {}
    for key, val in (calculated or {}).items():
        if key not in output_params:
            continue
        meta = output_params[key] or {}
        group_name = str(meta.get("group_name", "") or "").strip()

        if cfg.groups is not None and group_name in cfg.groups and not cfg.groups[group_name]:
            continue

        visible = _default_visible(meta)
        if cfg.fields is not None and key in cfg.fields:
            visible = bool(cfg.fields[key])
        if not visible:
            continue

        if _is_empty(val):
            continue

        grouped.setdefault(group_name, []).append(key)

    group_names = sorted(grouped.keys(), key=lambda s: s.lower())

    suffix_re = re.compile(cfg.suffix_regex)

    _default_tags = ["n", "out", "k"]
    allowed_tags = set(t.lower() for t in (cfg.transpose_tags or _default_tags))
    allowed_bases = set(str(b) for b in cfg.transpose_bases) if cfg.transpose_bases is not None else None

    def _strip_series_idx_from_header(header: str) -> str:
        s = (header or "").strip()
        s = re.sub(r"(?:\s|\()N\s*\d+\)?\s*$", "", s, flags=re.IGNORECASE).strip()
        return s

    for gname in group_names:
        keys = grouped[gname]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        set_cell(row, 1, gname if gname else "Без группы", font=bold, fill=group_fill, border_=border)
        row += 1

        def _sort_key(k: str):
            meta = output_params.get(k, {}) or {}
            return str(meta.get("header", k)).lower()

        keys_sorted = sorted(keys, key=_sort_key)

        suffix_series: dict[str, dict[str, dict[int, str]]] = {}
        suffix_base_header: dict[str, dict[str, str]] = {}
        freq_series: dict[str, dict[float, str]] = {}
        freq_base_header: dict[str, str] = {}
        freq_base_dim: dict[str, str] = {}

        normal: list[str] = []

        for k in keys_sorted:
            meta_k = output_params.get(k, {}) or {}
            header_k = str(meta_k.get("header", k))
            comment_k = str(meta_k.get("comment") or "")

            if not cfg.transpose_enabled:
                normal.append(k)
                continue

            if cfg.transpose_groups is not None and gname in cfg.transpose_groups and not cfg.transpose_groups[gname]:
                normal.append(k)
                continue

            if cfg.transpose_mode in ("auto", "freq"):
                fval = _extract_freq(header_k, comment_k)
                if fval is not None:
                    base_header = _strip_freq_suffix(header_k) or header_k
                    base = f"F__{_slug_base(base_header)}"
                    if allowed_bases is not None and base not in allowed_bases:
                        normal.append(k)
                        continue
                    base_used = base
                    base_header_used = base_header
                    freq_series.setdefault(base_used, {})[fval] = k
                    freq_base_header[base_used] = base_header_used
                    freq_base_dim.setdefault(base_used, str(meta_k.get("dimension", "") or ""))
                    continue

            if cfg.transpose_mode in ("auto", "suffix"):
                mk = suffix_re.match(str(k))
                if mk:
                    tag = str(mk.group("tag")).lower()
                    if allowed_tags and tag not in allowed_tags:
                        normal.append(k)
                        continue

                    base = str(mk.group("base"))
                    try:
                        idx = int(mk.group("idx"))
                    except Exception:
                        normal.append(k)
                        continue

                    if allowed_bases is not None and base not in allowed_bases:
                        normal.append(k)
                        continue

                    base_header = _strip_series_idx_from_header(header_k)
                    suffix_series.setdefault(tag, {}).setdefault(base, {})[idx] = k
                    suffix_base_header.setdefault(tag, {}).setdefault(base, base_header or base)
                    continue

            normal.append(k)

        for k in normal:
            meta = output_params.get(k, {}) or {}
            header = meta.get("header", k)
            dim = meta.get("dimension", "")
            comment = meta.get("comment") or ""

            v = calculated.get(k)
            f = _try_float(v)
            if f is not None and "accuracy" in meta:
                try:
                    f = round(f, int(meta.get("accuracy") or 0))
                except Exception:
                    pass
                v_out = f
            else:
                v_out = v

            set_cell(row, 1, header, border_=border)
            value_cell = set_cell(row, 2, v_out, border_=border)
            set_cell(row, 3, dim, border_=border)
            set_cell(row, 4, comment, border_=border)

            _push_debug(key=k, group=gname, meta=meta, value_cell_addr=value_cell.coordinate)

            row += 1

        for tag, base_map in suffix_series.items():
            if not base_map:
                continue

            if normal:
                row += 1

            idxs = sorted({idx for bm in base_map.values() for idx in bm.keys()})

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            set_cell(row, 1, f"_{tag}{{N}}", font=bold, fill=header_fill)
            row += 1

            set_cell(row, 1, "Параметр", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            for j, idx in enumerate(idxs, start=2):
                set_cell(row, j, f"{tag}{idx}", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            set_cell(row, 2 + len(idxs), "Ед.изм", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            row += 1

            def _base_sort_key(x):
                b = x[0]
                return (suffix_base_header.get(tag, {}).get(b, b) or b).lower()

            for base, idx_map in sorted(base_map.items(), key=_base_sort_key):
                first_key = idx_map[sorted(idx_map.keys())[0]]
                meta0 = output_params.get(first_key, {}) or {}
                header = suffix_base_header.get(tag, {}).get(base) or meta0.get("header", base)
                header = _strip_series_idx_from_header(str(header))
                dim = meta0.get("dimension", "")

                set_cell(row, 1, header, border_=border)

                for j, idx in enumerate(idxs, start=2):
                    k = idx_map.get(idx)
                    if not k:
                        set_cell(row, j, "", border_=border)
                        continue
                    meta_k = output_params.get(k, {}) or meta0
                    v = calculated.get(k)
                    f = _try_float(v)
                    if f is not None and "accuracy" in meta_k:
                        try:
                            f = round(f, int(meta_k.get("accuracy") or 0))
                        except Exception:
                            pass
                        cell = set_cell(row, j, f, border_=border)
                    else:
                        cell = set_cell(row, j, v, border_=border)

                    _push_debug(key=k, group=gname, meta=meta_k, value_cell_addr=cell.coordinate)

                set_cell(row, 2 + len(idxs), dim, border_=border)
                row += 1

            row += 1

        if freq_series:
            row += 1 if (normal or suffix_series) else 0

            freqs = sorted({f for m in freq_series.values() for f in m.keys()})

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            set_cell(row, 1, "Серии по частоте (Гц/Hz)", font=bold, fill=header_fill)
            row += 1

            def _fmt_f(f: float) -> str:
                if abs(f - int(f)) < 1e-9:
                    return str(int(f))
                s = f"{f}".replace(".", ",")
                return s

            set_cell(row, 1, "Параметр", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            for j, f in enumerate(freqs, start=2):
                set_cell(row, j, _fmt_f(f), font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            set_cell(row, 2 + len(freqs), "Ед.изм", font=bold, fill=header_fill, align=Alignment(horizontal="center"), border_=border)
            row += 1

            for base in sorted(freq_series.keys(), key=lambda b: (freq_base_header.get(b, b) or b).lower()):
                hdr = freq_base_header.get(base, base)
                dim = freq_base_dim.get(base, "")
                set_cell(row, 1, hdr, border_=border)

                idx_map = freq_series[base]
                for j, f in enumerate(freqs, start=2):
                    k = idx_map.get(f)
                    if not k:
                        set_cell(row, j, "", border_=border)
                        continue
                    meta_k = output_params.get(k, {}) or {}
                    v = calculated.get(k)
                    fv = _try_float(v)
                    if fv is not None and "accuracy" in meta_k:
                        try:
                            fv = round(fv, int(meta_k.get("accuracy") or 0))
                        except Exception:
                            pass
                        cell = set_cell(row, j, fv, border_=border)
                    else:
                        cell = set_cell(row, j, v, border_=border)

                    _push_debug(key=k, group=gname, meta=meta_k, value_cell_addr=cell.coordinate)

                set_cell(row, 2 + len(freqs), dim, border_=border)
                row += 1

            row += 1

        row += 1

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 60)

    ws.freeze_panes = "A5"

    if cfg.debug_sheet and debug_entries:
        ws_dbg = wb.create_sheet("DEBUG", 1)
        ws_dbg["A1"] = "DEBUG: значения (ссылки на лист 1)"
        ws_dbg["A1"].font = title_font
        ws_dbg.freeze_panes = "A2"
        ws_dbg.column_dimensions["A"].width = 34

        author = "dbg"
        r = 2
        for ent in debug_entries:
            addr = ent["addr"]
            sheet_title = ws.title
            formula = f"='{sheet_title}'!{addr}"
            cell = ws_dbg.cell(row=r, column=1, value=formula)
            cell.hyperlink = f"#'{sheet_title}'!{addr}"
            txt = (
                f"key: {ent['key']}\n"
                f"group: {ent['group']}\n"
                f"header: {ent['header']}\n"
                f"comment: {ent['comment']}\n"
                f"cell: {sheet_title}!{addr}"
            )
            try:
                cell.comment = Comment(txt, author)
            except Exception:
                pass
            r += 1

    if errors:
        ws_err = wb.create_sheet("Ошибки")
        ws_err["A1"] = "Параметры с ошибками расчёта"
        ws_err["A1"].font = title_font
        ws_err.append(["Параметр", "Сообщение"])
        ws_err["A2"].font = bold
        ws_err["B2"].font = bold
        for err in errors:
            ws_err.append([err.get("param", ""), err.get("msg", "")])
        ws_err.column_dimensions["A"].width = 45
        ws_err.column_dimensions["B"].width = 90

    try:
        wb.save(path)
    except Exception:
        return False

    return path


def sort_key_params(name: str, params: dict):
    trail_int_re = re.compile(r"(?:_|^)(\d+)$")
    meta = params.get(name, {}) or {}
    h = str(meta.get("header", name))
    c = str(meta.get("comment") or "")
    base_h = _strip_freq_suffix(h).lower()
    f = _extract_freq(h, c)
    m = trail_int_re.search(str(name))
    idx = int(m.group(1)) if m else None
    return (
        base_h,
        0 if f is not None else 1,
        f if f is not None else 0.0,
        idx if idx is not None else 10**18,
        str(name).lower(),
    )
