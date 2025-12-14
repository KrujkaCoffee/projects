import gspread
import project_cust_38.Cust_Functions as F
from openpyxl.utils import get_column_letter
try:
    import truststore
    truststore.inject_into_ssl()
except:
    print('truststore не загружен, нужна версия Пайтон 3.10. Работка googlesheets отключена')

keybook = '1NsTu-Kt9H83mPHiyWyC4lbHIbcrrp4V4vYmx5PKm_94'

class Gbook():
    FILE_KEY = r'Z:\Data\MKart\data\gold-rope-391308-285b8002ae52.json'
    def __init__(self,key_book):
        conn = gspread.service_account(filename=Gbook.FILE_KEY)
        # Указываем путь к JSON
        # Открываем тестовую book
        self.book = conn.open_by_key(key_book)

    def list_sheets(self):
        return [_.title for _ in self.book.worksheets()]


class Gsheet():
    def __init__(self,book:Gbook,name_sheet):
        # Открываем тестовую book
        self.sheet = book.book.worksheet(name_sheet)

    def cell_val(self,row_letter,column=0):
        if type(row_letter) == str:
            return self.sheet.acell(row_letter).value
        else:
            return self.sheet.cell(row_letter, column).value

    def cel_valrc(self,row,cell):
        letter = Gsheet._to_a1(row,cell)
        return Gsheet.cell_val(self,letter)
    
    @staticmethod
    def _to_a1(row, column):
        return get_column_letter(column) + str(row)

    @staticmethod
    def _to_r1c1(letter):
        number= ''
        letters = ''
        for item in letter:
            if F.is_numeric(item):
                number+= item
            else:
                letters += item

        num = F.alfabet_to_number(letters)
        return (num, int(number))

    def list_of_lists(self):
        return self.sheet.get_all_values()

    def list_of_dicts(self):
        return self.sheet.get_all_records()

    def cell_set(self,row_letter,column=0, value = ''):
        if type(row_letter) == str:
            self.sheet.update(row_letter, value)
        else:
            self.sheet.update_cell(row_letter, column, value)

if __name__ == '__main__':
    book = Gbook(key_book=keybook)
    list_sheets = book.list_sheets()
    plan_name = list_sheets[-1]
    sht = Gsheet(book,plan_name)
    print(sht.cell_val(4,1))
    #list_of_lists = sht.list_of_lists()
    sht.cell_set('Y49',value= F.now())
    print(sht.cell_val('Y49'))