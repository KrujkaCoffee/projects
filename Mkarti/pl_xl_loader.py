import re
from collections import defaultdict
from datetime import datetime, timedelta

from PyQt5 import QtGui

import dateutil
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidget, QComboBox
from openpyxl import load_workbook

from project_cust_38 import Cust_Qt as CQT
from project_cust_38 import Cust_SQLite as CSQ
from project_cust_38 import Cust_Functions as F
from project_cust_38 import Cust_mes as CMS
from project_cust_38 import Cust_progressBar as CPROG


class ExcelParser:
    def __init__(self, path) -> None:
        self.stop_column_index = None
        self.stop_row_index = None
        self.path = path
        self.get_need_column_index = 2
        self.need_time_index = 41
        self.absence_type = 55  # колонка с названием неявки
        self.absence_time = 57  # колонка с часами неявки

        self.table_name = 'сводная таблица'
        self.result = {}
        self.depatment = None
        self.month = None
        self.is_plan = None

    def get_start_stop_rows(self):
        self.max_row = self.ws.max_row
        self.start_row_index = + 2
        self.stop_row_index = self.max_row
        return True

    def load_page(self):
        sub_res = self.preparate_excel()
        if sub_res == 1:
            return
        if self.table_name in self.sheet_names:
            w = self.wb.sheetnames[1]
            if len(self.sheet_names) > 1:
                self.wb.remove_sheet(self.wb[w])
        return True

    def get_depatment_name(self):
        dep = self.ws.cell(row=8, column=2).value
        self.depatment = dep.lower()

    def get_month(self):
        row, col = self.find_cell('Отчетный период')
        month = self.ws.cell(row=row + 2, column=col).value
        if month:
            if isinstance(month, str):
                month = datetime.strptime(month, "%d.%m.%Y")
            self.month = month
            return True

    def preparate_excel(self):
        if self.get_start_stop_rows():
            self.get_position()
            self.get_depatment_name()
            return 0
        else:
            return 1

    def get_work_time(self, num_row, name):
        pre_time = self.ws.cell(row=num_row + 2, column=self.need_time_index).value
        if pre_time:
            prepared_value = str(pre_time).strip().replace(",", ".")
            self.add_in_inner_dict(name, 'Часы', float(prepared_value))

    def add_in_inner_dict(self, name, key, float_val):
        if not self.result.get(name):
            inner_di = {}
            inner_di[key] = float_val
            self.result[name] = inner_di
        else:
            if self.result[name].get(key):
                self.result[name][key] += float_val
            else:
                self.result[name][key] = float_val

    def get_absence(self, num_row, name, is_additional=False):
        sum_num = 0
        if is_additional:
            sum_num = 4
        for i in range(4):
            absence_type = self.ws.cell(row=num_row + i, column=self.absence_type + sum_num).value
            absence_time = self.ws.cell(row=num_row + i, column=self.absence_time + sum_num).value
            if absence_type and absence_time:
                absence_time_compile = re.compile(r'\(.\d*\)')
                absence_time_minutes = absence_time_compile.search(absence_time)
                if absence_time_minutes:
                    absence_time_minutes = absence_time_minutes.group()
                    self.add_in_inner_dict(name, absence_type, int(absence_time_minutes[1:-1]))

    def get_position(self):
        for num_row in range(self.start_row_index, self.stop_row_index):
            row = self.ws[num_row]
            value = row[self.get_need_column_index].value
            if value:
                find = re.search(r'\(.*\)', value)
                if find:
                    name = find[0][1:-1]
                    self.get_work_time(num_row, name)
                    self.get_absence(num_row, name)
                    self.get_absence(num_row, name, is_additional=True)

    def check_file_data(self):
        self.wb = load_workbook(self.path)
        self.sheet_names = self.wb.sheetnames
        sheet_name = self.wb.sheetnames[0]
        self.ws = self.wb[sheet_name]
        if not self.find_cell("рабочего времени"):
            return True

    def find_cell(self, find_val):
        for row in range(1, self.ws.max_row + 1):
            for col in range(1, self.ws.max_column + 1):
                val = self.ws.cell(row=row, column=col).value
                if val and isinstance(val, str) and find_val.strip().lower() in val.lower():
                    return row, col

    def check_file(self):
        msg = None
        if not self.path.endswith('.xlsx'):
            msg = 'Выбран не верный файл'
        elif self.check_file_data() or not self.get_month():
            msg = 'Выбран не верный файл'
        if msg:
            CQT.msgbox(msg)
            raise Exception(msg)

    def run(self, hook_prog_bar):
        try:
            hook_prog_bar.set(20)
            hook_prog_bar.text('Проверка документа')
            if err_msg := self.check_file():
                return err_msg
            hook_prog_bar.set(30)
            hook_prog_bar.text('Распаковка документа')
            if self.load_page():
                self.save_result()
                return True
        except PermissionError as e:
            CQT.msgbox('Перед выгрузкой закройте Excel!')
            return
        except Exception as e:
            CQT.msgbox('Произошла ошибка во время выгрузки excel')
            return

    def get_name_headers(self):
        all_times = set()
        for position_di in self.result.values():
            all_times.update(set(position_di.keys()))
        all_times.discard("Часы")
        li = ["Должность", "Часы", "Минуты"]
        li.extend(list(all_times))
        return li

    def get_table_headers(self):
        colunmn_name = self.get_name_headers()
        head_index = {}
        for num, name in enumerate(colunmn_name, start=1):
            self.save_ws.cell(column=num, row=1, value=name)
            head_index[name] = num
        return head_index

    def save_result(self):
        self.save_ws = self.wb.create_sheet(title=self.table_name)
        head_index = self.get_table_headers()
        del head_index["Должность"]
        row_num = 2
        for position, di in self.result.items():
            self.save_ws.cell(column=1, row=row_num, value=position)
            for name, value in di.items():
                self.save_ws.cell(column=head_index[name], row=row_num, value=value)
                if name == 'Часы':
                    self.save_ws.cell(column=head_index['Минуты'], row=row_num, value=value * 60)
            row_num += 1
        self.wb.save(self.path)


class PushWorkPlan:
    def __init__(
            self,
            window,
            plan_tbl,
            plan_tbl_fltr,
            fact_tbl,
            fact_tbl_fltr
    ):
        self.window = window
        self.buffered = None
        self.date_format = '%Y-%m-%d'
        self.bd_users = F.scfg('BD_users')
        self.bd_kplan = F.scfg('DB_kplan')
        self.naryad = F.scfg('Naryad')

        self.init_places()
        self.plan_tbl = plan_tbl
        self.plan_tbl_fltr = plan_tbl_fltr
        self.fact_tbl = fact_tbl
        self.fact_tbl_fltr = fact_tbl_fltr

        self.PODRAZDEL = F.deploy_dict_c(
            CSQ.custom_request_c(self.bd_kplan, f'SELECT Цвет, Наименование_ЕРП, Имя FROM podrazdel WHERE poki == {self.selected_poki} or poki is NULL', rez_dict=True),
            'Наименование_ЕРП'
        )

    @property
    def work_type_dep_lis(self):
        return self.get_vid_dep_list()

    @property
    def name_worktypes(self):
        return self.load_name_worktypes()

    @property
    def selected_place(self):
        current_poki = self.window.ui.cmb_pl_tabel_place.currentData()
        return self.PLACES_BY_POKI[current_poki]

    @property
    def selected_poki(self):
        return self.window.ui.cmb_pl_tabel_place.currentData()

    @property
    def months(self):
        return sorted(self.generate_months())

    def get_vid_dep_list(self):
        code_erp = self.selected_place.get('РодительВидаРабот')
        result = CSQ.custom_request_c(
            self.bd_users,
            f'''
            SELECT 
                vrpd.Вид_работ, 
                p.Наименование_ЕРП
            FROM vid_rab_po_dolg AS vrpd
            INNER JOIN podrazdel AS p ON p.Имя = vrpd.name_tbl
            WHERE vrpd.Родитель = {code_erp!r}
        ''', hat_c=False, attach_dbs=self.bd_kplan)
        if result == False:
            return CQT.msgbox('Не удалось запросить данные о видах работ и цехах')
        return [
            '\n'.join((vid, dep.lower()))
            for vid, dep in result
        ]

    def get_list_places(self):
        response = CSQ.custom_request_c(
            self.naryad,
            'SELECT * FROM places',
            rez_dict=True
        )
        if not isinstance(response, list):
            CQT.msgbox('Не удалось загрузить организации')
            return []
        return response

    def is_plan(self, err_msg: str = 'Штат->Планы\nНе вычислена таблица БД т.к. вкладка не распознана'):
        sub_tab_ind = self.window.ui.tabWidget_10.currentIndex()
        tab_text = self.window.ui.tabWidget_10.tabText(sub_tab_ind)
        if tab_text == 'План':
            return True
        if tab_text == 'Факт':
            return False
        raise Exception(err_msg)

    @property
    def table_name(self):
        if self.is_plan('Штат->Планы\nНе вычислена таблица БД т.к. вкладка не распознана'):
            return 'plan_tabel_workforce'
        return 'fact_tabel_workforce'

    @property
    def table_widget(self):
        if self.is_plan('Штат->Планы\nНе вычислен виджет таблицы т.к. вкладка не распознана'):
            return self.plan_tbl
        return self.fact_tbl

    @property
    def table_filter(self):
        if self.is_plan('Штат->Планы\nНе вычислен виджет таблицы т.к. вкладка не распознана'):
            return self.plan_tbl_fltr
        return self.fact_tbl_fltr

    @property
    def first_day_cur_month(self):
        date = datetime.now().replace(day=1)
        if not self.is_plan():
            date = date - dateutil.relativedelta.relativedelta(months=2)
        return date

    def load_name_worktypes(self):
        stmt = f"""
            SELECT имя, вид_работ 
            FROM professions 
            WHERE примечание != "не исп" AND примечание != "не исп, старый" AND poki = {self.selected_poki}
        """
        work_types = F.deploy_dict_c(CSQ.custom_request_c(F.scfg('BD_users'), stmt, rez_dict=True), 'имя')
        if not work_types:
            CQT.msgbox('Не удалось загрузить виды работ')
            return
        return work_types

    def remove_month(self):
        if not CMS.user_access(self.naryad, 'мкарт_внесение_корректировка_планов_фактов', F.user_name()):
            return
        current_row = self.table_widget.currentRow()
        if current_row == -1:
            return
        name_month = self.table_widget.verticalHeaderItem(current_row).text()
        if not CQT.msgboxgYN(f'Вы уверены что хотите удалить месяц: {name_month}'):
            return
        response = CSQ.custom_request_c(
            self.bd_kplan,
            f"""
                DELETE FROM {self.table_name}
                WHERE month = "{name_month} AND poki = {self.selected_poki}"
            """
        )
        msg = 'Не удалось удалить месяц. Попробуйте ещё раз'
        if response:
            self.fill_tbl()
            CQT.fill_filtr_c(self, self.table_filter, self.table_widget)
            CMS.apply_filtr_c(self, self.table_filter, self.table_widget)
            msg = f'Месяц {name_month} успешно удален!'
        CQT.msgbox(msg)

    def copy_data(self):
        row = self.table_widget.currentRow()
        self.buffered = row
        text = self.table_widget.verticalHeaderItem(row).text()
        CQT.msgbox(f'Скопирован месяц: {text}')

    @CQT.onerror
    def delete_insert_vid_to_month(self, **kwargs):
        if all(value != '' for value in kwargs.values()):
            response = set()
            keys = ', '.join(kwargs.keys())
            values = ', '.join(f'{value!r}' for value in kwargs.values())
            vid_rabot = kwargs[
                'vid_rabot']  # через нотацию квадратных скобок для быстрого отказа в случае отсутствия ключа
            month = kwargs['month']
            poki = self.selected_poki
            instructions = (
                f'''DELETE FROM {self.table_name} WHERE vid_rabot = "{vid_rabot}" AND month = "{month}" AND poki = {poki}''',
                f'''INSERT INTO {self.table_name}({keys}) VALUES ({values})'''
            )
            for instruction in instructions:
                response.add(CSQ.custom_request_c(self.bd_kplan, instruction))
            return all(response) and len(response) == len(instructions)

    def get_last_months(self):
        date = self.first_day_cur_month
        data = CSQ.custom_request_c(self.bd_kplan, f'''
            SELECT month, vid_rabot, depatment, normo_smen
            FROM {self.table_name} 
            WHERE poki = {self.selected_poki} AND date(month) >= date("{date.strftime("%Y-%m-%d")}")
        ''', rez_dict=True)
        if data is None or data is False:
            CQT.msgbox('Не удалось загрузить данные для таблицы')
            return
        return data

    def last_month(self):
        last_row = self.table_widget.rowCount() - 1
        if self.table_widget.rowCount() == 0:
            return datetime.now().replace(day=1).strftime(self.date_format)
        last_months_on_tbl = self.table_widget.verticalHeaderItem(last_row).text()
        if F.is_date(last_months_on_tbl, self.date_format):
            date_obj = datetime.strptime(last_months_on_tbl, self.date_format) + dateutil.relativedelta.relativedelta(
                months=1)
            return date_obj.strftime(self.date_format)

    def insert_one_month(self):
        if not CMS.user_access(self.naryad, 'мкарт_внесение_корректировка_планов_фактов', F.user_name()):
            return
        month = self.last_month()
        if not CQT.msgboxgYN(f'Вы действительно хотите добавить месяц: {month}?'):
            return
        months = self.get_last_months()
        data = self.aggregate_months(months)
        if data:
            self.fill_tbl([*data, [''] * len(self.work_type_dep_lis)])

    def paste_data(self):
        if not CMS.user_access(self.naryad, 'мкарт_внесение_корректировка_планов_фактов', F.user_name()):
            return
        if self.buffered is None:
            CQT.msgbox('СНачала нужно скопировать строчку!')
            return
        row = self.table_widget.currentRow()
        name_month = self.table_widget.verticalHeaderItem(row).text()
        if self.buffered == row:
            return
        for col in range(self.table_widget.columnCount()):
            copy_text = self.table_widget.item(self.buffered, col).text()
            self.table_widget.item(row, col).setText(copy_text)
            vert_head = self.table_widget.horizontalHeaderItem(col).toolTip()
            vid_rabot, depatment = vert_head.split('\n')
            self.delete_insert_vid_to_month(vid_rabot=vid_rabot, month=name_month, normo_smen=copy_text,
                                            depatment=depatment, poki=self.selected_poki)

    def generate_months(self):
        dates = []
        current_date = self.first_day_cur_month
        for i in range(16):
            new_date = current_date + dateutil.relativedelta.relativedelta(months=i)  # Добавляем i месяцев
            dates.append(new_date.strftime('%Y-%m-%d'))
        return sorted(dates)

    def aggregate_months(self, data: list):
        agg_months = {}
        start_month = self.first_day_cur_month.strftime("%Y-%m-%d")
        for row in data:
            if month := row.get('month'):
                key = f'{row.get("vid_rabot")}\n{row.get("depatment")}'
                if month >= start_month:
                    agg_months.setdefault(month, {})[key] = row.get('normo_smen')

        agg_vid = [self.work_type_dep_lis]
        for month in self.months:
            tmp = []
            for vid in agg_vid[0]:
                cur_month_data = agg_months.get(month, {})
                if normo_smen := cur_month_data.get(vid):
                    tmp.append(normo_smen)
                else:
                    tmp.append('')
            agg_vid.append(tmp)
        return agg_vid

    def get_department_data(self, key: str):
        for dep_erp in self.PODRAZDEL:
            if dep_erp.lower() == key.lower():
                return self.PODRAZDEL[dep_erp]

    def decor_table(self):
        col_count = self.table_widget.columnCount()
        for col in range(col_count):
            head_item = self.table_widget.horizontalHeaderItem(col)
            full_name = head_item.text()
            head_item.setToolTip(full_name)
            vid, dep = full_name.split('\n')
            if data := self.get_department_data(dep):
                r, g, b = data.get('Цвет', '255;255;255').split(';')
                name = data.get('Имя', '').replace('пл_', '')
                head_item.setText(f'{vid}({name})')
                color_inst = QtGui.QColor(int(r), int(g), int(b))
                head_item.setForeground(color_inst)
                color_inst = QtGui.QColor(int(r), int(g), int(b), 32)
                font = head_item.font()
                font.setBold(True)
                head_item.setFont(font)
                for row in range(self.table_widget.rowCount()):
                    item = self.table_widget.item(row, col)
                    item.setBackground(color_inst)
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setToolTip('нормо-смен')
        #self.table_widget.horizontalHeader().setFixedHeight(40)

    @CQT.onerror
    def fill_tbl(self, data=None):
        self.table_widget.setRowCount(0) or self.table_widget.setColumnCount(0)
        if data is None:
            months = self.get_last_months()
            data = self.aggregate_months(months)
        CQT.fill_wtabl(data, self.table_widget, height_row=24, min_width_col=0, ogr_maxshir_kol=500,
                       StretchLastSection=False,list_column_widths=CMS.load_column_widths(self,self.table_widget))
        self.table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_widget.setVerticalHeaderLabels(self.months)
        self.table_widget.verticalHeader().setStyleSheet("""
            QHeaderView::section {
                border: 1px solid #A9A9A9;
                padding: 5px;
                font-size: 14px;
            }
        """)
        self.decor_table()
        CQT.fill_filtr_c(self, self.table_filter, self.table_widget, hidden_scroll=True)
        CMS.update_width_filtr(self.table_widget, self.table_filter)
        CMS.apply_filtr_c(self, self.table_filter, self.table_widget)

    def accum_work_time(self, data):
        agr_vid_rabot = defaultdict(float)
        for position_name in data.keys():
            vid_rabot = self.name_worktypes.get(position_name)
            if vid_rabot:
                agr_vid_rabot[vid_rabot] += data[position_name]['Часы']
        return agr_vid_rabot

    @CPROG.progress_decorator
    def load_xlsx(self, hook_prog_bar):
        if not CMS.user_access(self.naryad, 'мкарт_внесение_корректировка_планов_фактов', F.user_name()):
            return
        tmp_path = CMS.load_tmp_path('btn_load_plan_fact_xl')
        path = CQT.f_dialog_name(self.window, 'Выбрать файл', tmp_path, f"Файлы (*xlsx)")
        if path is None or path == '.':
            return
        hook_prog_bar.set(10)
        hook_prog_bar.text('Инициализация excel')
        ep = ExcelParser(path)
        CMS.save_tmp_path('btn_load_plan_fact_xl', path, True)
        if ep.run(hook_prog_bar) is not None:
            hook_prog_bar.set(70)
            hook_prog_bar.text('Сохранение результатов в БД')
            month = ep.month.strftime("%Y-%m-%d")
            work_time = self.accum_work_time(ep.result)
            for wt, time in work_time.items():
                if time != 0:
                    normo_smen = round(time / 8, 2)
                    self.delete_insert_vid_to_month(
                        depatment=ep.depatment,
                        vid_rabot=wt,
                        month=month,
                        normo_smen=normo_smen,
                        poki=self.selected_poki
                    )
            hook_prog_bar.set(90)
            hook_prog_bar.text('Заполнение таблицы актуальными данными')
            self.table_widget.clear()
            self.fill_tbl()
            CQT.msgbox(f'Месяц {month} успешно выгружен')

    def init_places(self):
        self.PLACES_BY_POKI = F.deploy_dict_c(self.get_list_places(), 'poki')
        cmb: QComboBox = self.window.ui.cmb_pl_tabel_place
        for poki, place in self.PLACES_BY_POKI.items():
            cmb.addItem(place['Имя'], poki)
        # # TODO блок смены организации
        #cmb.setCurrentIndex(0)
        #cmb.setDisabled(True)
