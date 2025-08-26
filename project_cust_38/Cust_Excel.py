import copy
from builtins import range
from typing import Generator, Iterable, Any
from datetime import datetime
import xlwings as XL
import project_cust_38.Cust_Functions as F
import openpyxl as xlo
from openpyxl.utils import get_column_letter
from project_cust_38.Cust_progressBar import progress_decorator as CQTPD
try:
    from docxtpl import DocxTemplate, RichText
except Exception as e:
    print(e)

try:
    from PyQt5 import   QtGui, QtCore ,QtWidgets
except:
    pass
#https://programmersought.com/article/81673784973/

BORDER_DICT = {'xlEdgeLeftAll': 1,  # Left edge of each cell in range, not in enumeration docs
                       'xlEdgeRightAll': 2,  # Right edge of each cell in range, not in enumeration docs
                       'xlEdgeTopAll': 3,  # Top edge of each cell in range, not in enumeration docs
                       'xlEdgeBottomAll': 4,  # Bottom edge of each cell in range, not in enumeration docs
                       'xlDiagonalDown': 5,
                       # Border running from the upper-left corner to the lower-right of each cell in the range.
                       'xlDiagonalUp': 6,
                       # Border running from the lower-left corner to the upper-right of each cell in the range.
                       'xlEdgeLeft': 7,  # Border at the left edge of the range.
                       'xlEdgeTop': 8,  # Border at the top of the range.
                       'xlEdgeBottom': 9,  # Border at the bottom of the range.
                       'xlEdgeRight': 10,  # Border at the right edge of the range.
                       'xlInsideVertical': 11,
                       # Vertical borders for all the cells in the range except borders on the outside of the range.
                       'xlInsideHorizontal': 12}  # Horizontal borders for all cells in the range except borders on the outside of the range.



def spis_listov(wb):
    rez = []
    for i in range(0,wb.sheets.count):
        rez.append(wb.sheets[i].name)
    return rez

def read_file(book,list_name,r1=1,r2='*',c1=1,c2='*'):
    def lastRC_xlo(ws):
        max_col_row = 0
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            col_row = len([cell for cell in ws[col_letter] if cell.value])
            if max_col_row < col_row:
                max_col_row = col_row
        return [max_col_row, ws.max_column]

    def RangeValue(sheet, startRow, endRow, startCol, endCol ):
        rangeSelected = []
        # Loops through selected Rows
        for i in range(startRow, endRow + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol, endCol + 1, 1):
                rowSelected.append(sheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected
    
    workbook = xlo.load_workbook(book)
    worksheet = workbook.get_sheet_by_name(list_name)

    r,c = lastRC_xlo(worksheet)
    if r2 == '*':
        r2 = r 
    if c2 == '*':
        c2 = c 
    rez = RangeValue(worksheet, r1,r2,c1,c2)
    #XL.App().visible = True

    return rez





def lastRow(sheet, col=1):
    """ Find the last row in the worksheet that contains data.

    idx: Specifies the worksheet to select. Starts counting from zero.

    workbook: Specifies the workbook

    col: The column in which to look for the last cell containing data.
    """

    lwr_r_cell = sheet.cells.last_cell      # lower right cell
    lwr_row = lwr_r_cell.row             # row of the lower right cell
    lwr_cell = sheet.range((lwr_row, col))  # change to your specified column

    if lwr_cell.value is None:
        lwr_cell = lwr_cell.end('up')    # go up untill you hit a non-empty cell

    return lwr_cell.row

def ogran(spis):
    if spis == [[]]:
        return ['']
    for i in range(len(spis)):
        for j in range(len(spis[i])):
            if type(spis[i][j]) != type('1') and type(spis[i][j]) != type(1) : 
                spis[i][j] = str(spis[i][j])
            if type(spis[i][j]) == type("1"):
                if spis[i][j].count('\n') > 200:
                    spis[i][j] = spis[i][j].replace('\n', ' ')
                if len(spis[i][j]) > 320:
                    spis[i][j] = spis[i][j][:320]
                if spis[i][j] != '':
                    if spis[i][j][0] == '=':
                        spis[i][j] = "'" + spis[i][j]
    return spis

def zap_spis(spisok,putf,wb_name,ws_name,row,column,autofit=True, zag_bold=True,orient_g_v = 'v',return_putf=False):
    spisok = ogran(spisok)
    if '.xlsx' not in wb_name:
        raise ValueError
    if not F.existence_file_c(putf):
        F.create_dir_c(putf)
    file_path = putf + F.sep() + F.clear_row_for_file_name_c(wb_name)
    if F.existence_file_c(file_path):
        try:
            F.delete_file_c(file_path)
        except:
            print('Файл занят')
            return False
    try:

        app = XL.App(visible=False, add_book=True)
        #if F.existence_file_c(file_path) == False:
        #    #wb = app.books.add()
        #    #wb.save(path=file_path)
        #    #wb.close()
        wb = app.books.add()
        if ws_name not in spis_listov(wb):
           wb.sheets.add(name=ws_name)
        sheet = wb.sheets[ws_name]
        sheet[row, column].options(expand='table').value = spisok
        rng = sheet[row:row + len(spisok), column:column + len(spisok[0])]
        rng.api.Borders.LineStyle = 1
        if zag_bold:
           rng2 = sheet[row:row +1, column:column+len(spisok[0])]
           rng2.api.Font.Bold = True
    
        if autofit:
           for ws in wb.sheets:
               ws.autofit(axis="columns")
        if orient_g_v == 'v':
           sheet.page_setup.api._inner.Orientation = 1.0
        else:
           sheet.page_setup.api._inner.Orientation = 2.0
        wb.save(path=file_path)
        wb.close()
        app.kill()
        rez = True

    except:
        print('Ошибка экспорта в ексель')
        rez = False
    finally:
        try:
            wb.close()
        except:
            pass
        try:
            app.kill()
        except:
            pass
        if return_putf and rez == True:
            return file_path
        return rez
    
def pechat_table(spisok,row,column, orient_g_v = 'v', zag_bold = True, otstup_l=5.0, otstup_r=5.0,otstup_v=5.0,otstup_n=5.0,zoom=100,copies=1):
    #file_path = F.put_po_umolch() + F.sep() + "wb_tmp"
    app = XL.App(visible=False, add_book=False)
    wb = app.books.add()
    wb.sheets.add(name='ws_name')
    sheet = wb.sheets['ws_name']
    sheet[row, column].options(expand='table').value = spisok
    sheet[row:row+ len(spisok), column:column+len(spisok[0])].columns.autofit()
    rng = sheet[row:row+ len(spisok), column:column+len(spisok[0])]
    rng.api.Borders.LineStyle = 1

    if zag_bold:
        rng2 = sheet[row:row +1, column:column+len(spisok[0])]
        rng2.api.Font.Bold = True

    if orient_g_v == 'v':
        sheet.page_setup.api._inner.Orientation = 1.0
    else:
        sheet.page_setup.api._inner.Orientation = 2.0

    sheet.page_setup.api._inner.LeftMargin = otstup_l
    sheet.page_setup.api._inner.TopMargin = otstup_v
    sheet.page_setup.api._inner.RightMargin = otstup_r
    sheet.page_setup.api._inner.BottomMargin=otstup_n
    sheet.page_setup.api._inner.Zoom = zoom

    rng.api.PrintOut(Copies=copies)

    #wb.save()
    wb.close()
    app.quit()

@CQTPD
def save_table_colour(tbl,putf:str,wb_name_wout_exe:str,ws_name:str,row:int=1,column:int=1,hat:list=None,wo_hide_rows_cols=False,print_hat_tbl=True,hook_prog_bar=None):
    wb_name = wb_name_wout_exe + ".xlsx"
    file_path = putf + F.sep() + F.clear_row_for_file_name_c(wb_name)
    if F.existence_file_c(file_path):
        try:
            F.delete_file_c(file_path)
        except:
            print('Файл занят')
            return False
    
    app = XL.App(visible=False, add_book=True)
    #if F.existence_file_c(file_path) == False:
    #    #wb = app.books.add()
    #    #wb.save(path=file_path)
    #    #wb.close()
    wb = app.books.add()
    if ws_name not in spis_listov(wb):
       wb.sheets.add(name=ws_name)
    sheet = wb.sheets[ws_name]
    hook_prog_bar.set(0)
    hook_prog_bar.text('Обработка Заголовков')
    if isinstance(hat,list):
        row_tmp = copy.deepcopy(row)
        column_tmp = copy.deepcopy(column)
        for j in range(len(hat[0])):
            for i in range(len(hat)):
                font = sheet.range((row_tmp + i, column_tmp + j)).font
                font.size = 10
                font.bold = True
                val = hat[i][j]
                sheet.range((row_tmp + i, column_tmp + j)).value = val
                sheet.range((row_tmp + i, column_tmp + j)).api.WrapText = True
        row+=len(hat)+2

    row_tmp = copy.deepcopy(row)+1
    column_tmp = copy.deepcopy(column)
    if print_hat_tbl == False:
        row_tmp -= 1
    koef_hide = 0
    for i in range(tbl.rowCount()):
        hook_prog_bar.set(5 + round(i/tbl.rowCount()*20))
        hook_prog_bar.text('Обработка Горизонтальнго заголовка')
        rowHeight =int(tbl.rowHeight(i)*0.768)
        if rowHeight > 255:
            rowHeight = 255
        if wo_hide_rows_cols:  
            if tbl.isRowHidden(i) or tbl.rowHeight(i) <= 1:
                #sheet.range((row_tmp, column_tmp + j)).column_width = 0
                koef_hide +=1
                continue
        
        sheet.range((row_tmp + i-koef_hide, column_tmp)).row_height = rowHeight
        if tbl.verticalHeaderItem(i) == None:
            sheet.range((row_tmp + i-koef_hide, column_tmp)).value = ''

        else:
            name_head_row = tbl.verticalHeaderItem(i).text()
            item = tbl.verticalHeaderItem(i)
            font = item.font()
            col_obj = item.foreground()
            r, g, b, a = col_obj.color().getRgb()
            size = font.pointSize()
            bold = font.bold()

            sheet.range((row_tmp + i - koef_hide, column_tmp)).value = name_head_row
            sheet.range((row_tmp + i - koef_hide, column_tmp)).api.WrapText = True
            #sheet.Range((row_tmp + i, column_tmp)).color = (r, g, b)

            font = sheet.range((row_tmp + i-koef_hide, column_tmp)).font
            font.color = (r, g, b)
            font.size = size
            font.bold = bold

    row_tmp = copy.deepcopy(row)
    column_tmp = copy.deepcopy(column) + 1
    if tbl.verticalHeaderItem(0) == None:
        column_tmp -=1
    koef_hide = 0
    for j in range(tbl.columnCount()):
        hook_prog_bar.set(25 + round(j/tbl.columnCount()*20))
        hook_prog_bar.text('Обработка Вертиклаьного заголовка')
        if wo_hide_rows_cols:
            if tbl.isColumnHidden(j) or tbl.columnWidth(j) <= 1:
                #sheet.range((row_tmp, column_tmp + j)).column_width = 0
                koef_hide+=1
                continue

        column_width = px2ch(tbl.columnWidth(j))
        if column_width > 255:
            column_width = 255
        if column_width < 2:
            column_width = 2
        sheet.range((row_tmp, column_tmp + j-koef_hide)).column_width = column_width

        if tbl.horizontalHeaderItem(j) == None or print_hat_tbl==False:
            sheet.range((row_tmp, column_tmp + j-koef_hide)).value = ''


        else:
            name_head_col = tbl.horizontalHeaderItem(j).text()
            item = tbl.horizontalHeaderItem(j)
            font = item.font()
            col_obj = item.foreground()
            r, g, b, a = col_obj.color().getRgb()
            size = font.pointSize()
            bold = font.bold()
            sheet.range((row_tmp,column_tmp+j-koef_hide)).value = name_head_col
            sheet.range((row_tmp,column_tmp+j-koef_hide)).api.WrapText = True
            #sheet.Range((row_tmp + i, column_tmp)).color = (r, g, b)

            font = sheet.range((row_tmp, column_tmp+j-koef_hide)).font
            font.color = (r, g, b)
            font.size = size
            font.bold = bold

    row_tmp = copy.deepcopy(row) +1
    column_tmp = copy.deepcopy(column) + 1

    if print_hat_tbl == False:
        row_tmp -= 1
    if tbl.verticalHeaderItem(0) == None:
        column_tmp -=1
    column_counter = 0
    koef_hide_c =0
    for j in range(tbl.columnCount()):

        if wo_hide_rows_cols:
            if tbl.isColumnHidden(j) or tbl.columnWidth(j) <= 1:
                koef_hide_c+=1
                continue
        column_counter += 1
        row_counter = 0
        koef_hide = 0
        for i in range(tbl.rowCount()):
            hook_prog_bar.set(45 + round((tbl.rowCount()*j+i) / (tbl.columnCount()*tbl.rowCount()) * 50))
            hook_prog_bar.text('Обработка таблицы')
            if wo_hide_rows_cols:
                if tbl.isRowHidden(i) or tbl.rowHeight(i) <= 1:
                    koef_hide +=1
                    continue
            row_counter+=1
            if tbl.item(i, j) == None:
                #sheet.range((row_tmp+i-koef_hide, column_tmp + j)).value = ''
                pass
            else:
                item = tbl.item(i, j)
                font = item.font()
                col_obj = item.foreground()
                r, g, b, a = col_obj.color().getRgb()
                size = font.pointSize()
                bold = font.bold()
                font = sheet.range((row_tmp+i-koef_hide, column_tmp + j -koef_hide_c)).font
                font.color = (r, g, b)
                font.size = size
                font.bold = bold

                val = tbl.item(i, j).text()
                r, g, b, a =item.background().color().getRgb()
                if F.is_numeric(val):
                    if isinstance(F.valm(val) ,int):
                        try:
                            sheet.range((row_tmp+i-koef_hide, column_tmp + j)).number_format = '0'
                        except:
                            pass
                    else:
                        try:
                            sheet.range((row_tmp+i-koef_hide, column_tmp + j)).number_format = '0.00'
                        except:
                            pass
                else:
                    sheet.range((row_tmp+i-koef_hide, column_tmp + j)).number_format = "@"
                sheet.range((row_tmp+i-koef_hide, column_tmp + j-koef_hide_c)).value = val
                sheet.range((row_tmp+i-koef_hide, column_tmp + j-koef_hide_c)).api.WrapText = True
                if r==g==b==0:
                    pass
                else:
                    sheet.range((row_tmp+i-koef_hide, column_tmp + j-koef_hide_c)).color = (r, g, b)
                # sheet.Range((row_tmp + i, column_tmp)).color = (r, g, b)
    hook_prog_bar.set(100)
    hook_prog_bar.text('сохранение')
    row_tmp = copy.deepcopy(row) + 1
    column_tmp = copy.deepcopy(column) + 1
    if print_hat_tbl == False:
        row_tmp -= 1
    if tbl.verticalHeaderItem(0) == None:
        column_tmp -=1


    

    def set_border(rng,type,weigth,color):
        rng.api.Borders(type).Weight = weigth+1
        rng.api.Borders(type).Color = color


    rng = sheet.range((row_tmp, column_tmp), (row_tmp + row_counter - 1, column_tmp + column_counter - 1))


    if 'custBorderInfo' in tbl.__dict__:
        custBorderInfo = tbl.custBorderInfo
        for i in range(row_tmp-1, row_tmp + row_counter-1):
            for j in range(column_tmp-1, column_tmp + column_counter-1):
                rng = sheet.range((i+1, j+1))
                if custBorderInfo.thick_out > 0:
                    if (i,j) in custBorderInfo.filled_bottom:
                        set_border(rng, 9, custBorderInfo.thick_out, custBorderInfo.color_out.rgb())
                    if (i,j) in custBorderInfo.filled_left:
                        set_border(rng, 7, custBorderInfo.thick_out, custBorderInfo.color_out.rgb())
                    if (i,j) in custBorderInfo.filled_right:
                        set_border(rng, 10, custBorderInfo.thick_out, custBorderInfo.color_out.rgb())
                    if (i,j) in custBorderInfo.filled_top:
                        set_border(rng, 8, custBorderInfo.thick_out, custBorderInfo.color_out.rgb())
                if custBorderInfo.thick_in > 0:
                    if (i,j) in custBorderInfo.inside_right:
                        set_border(rng, 10, custBorderInfo.thick_in, custBorderInfo.color_in.rgb())
                    if (i,j) in custBorderInfo.inside_top:
                        set_border(rng, 8, custBorderInfo.thick_in, custBorderInfo.color_in.rgb())
    else:
        format_borders(rng, 2, color_str='133,133,133')

    wb.save(path=file_path)
    wb.close()
    app.kill()
    rez = file_path
    #except:
    #    print('Ошибка экспорта в ексель')
    #    rez = False
    #finally:
    #    try:
    #        wb.close()
    #    except:
    #        pass
    #    try:
    #        app.kill()
    #    except:
    #        pass
    #    return rez
    return rez


def save_table_colour_openpyxl(tbl, putf: str, wb_name_wout_exe: str, ws_name: str, row: int = 1, column: int = 1,
                               hat: list = None, wo_hide_rows_cols=False, print_hat_tbl=True, hook_prog_bar=None):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    wb_name = wb_name_wout_exe + ".xlsx"
    file_path = putf + F.sep() + F.clear_row_for_file_name_c(wb_name)
    print(file_path)
    if F.existence_file_c(file_path):
        try:
            F.delete_file_c(file_path)
        except:
            print('Файл занят')
            return False
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    if isinstance(hat, list):
        row_tmp = copy.deepcopy(row)
        column_tmp = copy.deepcopy(column)
        for j in range(len(hat[0])):
            for i in range(len(hat)):
                font = sheet.range((row_tmp + i, column_tmp + j)).font
                val = hat[i][j]

                cell = sheet.cell(row=row_tmp + i, column=column_tmp + j)
                cell.font = Font(
                    size=10,
                    bold=True,
                    name='Arial'
                )
                cell.value = val
                cell.alignment = Alignment(wrap_text=True, horizontal='center')
                # font.size = 10
                # font.bold = True
                # sheet.range((row_tmp + i, column_tmp + j)).value = val
                # sheet.range((row_tmp + i, column_tmp + j)).api.WrapText = True
        row += len(hat) + 2

    row_tmp = copy.deepcopy(row) + 1
    column_tmp = copy.deepcopy(column)
    if print_hat_tbl == False:
        row_tmp -= 1
    koef_hide = 0
    for i in range(tbl.rowCount()):
        rowHeight = int(tbl.rowHeight(i) * 0.768)
        if rowHeight > 255:
            rowHeight = 255
        if wo_hide_rows_cols:
            if tbl.isRowHidden(i) or tbl.rowHeight(i) <= 1:
                # sheet.range((row_tmp, column_tmp + j)).column_width = 0
                koef_hide += 1
                continue
        row_index = row_tmp + i - koef_hide
        sheet.row_dimensions[row_index].height = rowHeight
        # sheet.range((row_tmp + i - koef_hide, column_tmp)).row_height = rowHeight
        if tbl.verticalHeaderItem(i) == None:
            row_index = row_tmp + i - koef_hide
            sheet.cell(row=row_index, column=column_tmp, value='')
        else:
            name_head_row = tbl.verticalHeaderItem(i).text()
            item = tbl.verticalHeaderItem(i)
            font = item.font()
            col_obj = item.foreground()
            r, g, b, a = col_obj.color().getRgb()
            size = font.pointSize()
            bold = font.bold()
            cell = sheet.cell(row=row_tmp + i - koef_hide, column=column_tmp)
            cell.value = name_head_row
            cell.alignment = Alignment(wrap_text=True)
            font_color = f"{r:02X}{g:02X}{b:02X}"
            cell.font = Font(size=size, bold=bold, color=font_color, name='Arial')

    row_tmp = copy.deepcopy(row)
    column_tmp = copy.deepcopy(column) + 1
    if tbl.verticalHeaderItem(0) == None:
        column_tmp -= 1
    koef_hide = 0
    for j in range(tbl.columnCount()):
        if wo_hide_rows_cols:
            if tbl.isColumnHidden(j) or tbl.columnWidth(j) <= 1:
                # sheet.range((row_tmp, column_tmp + j)).column_width = 0
                koef_hide += 1
                continue

        column_width = px2ch(tbl.columnWidth(j))
        if column_width > 255:
            column_width = 255
        if column_width < 2:
            column_width = 2
        # sheet.range((row_tmp, column_tmp + j - koef_hide)).column_width = column_width

        column_index = column_tmp + j - koef_hide
        sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = column_width
        if tbl.horizontalHeaderItem(j) == None or print_hat_tbl == False:
            sheet.cell(row=row_tmp, column=column_tmp + j - koef_hide, value='')
        else:
            name_head_col = tbl.horizontalHeaderItem(j).text()
            item = tbl.horizontalHeaderItem(j)
            font = item.font()
            col_obj = item.foreground()
            r, g, b, a = col_obj.color().getRgb()
            size = font.pointSize()
            bold = font.bold()
            row_index = row_tmp
            column_index = column_tmp + j - koef_hide
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = name_head_col
            cell.alignment = Alignment(wrap_text=True)
            font_color = f"{r:02X}{g:02X}{b:02X}"
            cell.font = Font(color=font_color, size=size, bold=bold, name='Arial')

    row_tmp = copy.deepcopy(row) + 1
    column_tmp = copy.deepcopy(column) + 1

    if print_hat_tbl == False:
        row_tmp -= 1
    if tbl.verticalHeaderItem(0) == None:
        column_tmp -= 1
    column_counter = 0
    koef_hide_c = 0
    for j in range(tbl.columnCount()):

        if wo_hide_rows_cols:
            if tbl.isColumnHidden(j) or tbl.columnWidth(j) <= 1:
                koef_hide_c += 1
                continue
        column_counter += 1
        row_counter = 0
        koef_hide = 0
        for i in range(tbl.rowCount()):
            if wo_hide_rows_cols:
                if tbl.isRowHidden(i) or tbl.rowHeight(i) <= 1:
                    koef_hide += 1
                    continue
            row_counter += 1
            if tbl.item(i, j) == None:
                # sheet.range((row_tmp+i-koef_hide, column_tmp + j)).value = ''
                pass
            else:
                item = tbl.item(i, j)
                font = item.font()
                col_obj = item.foreground()
                r, g, b, a = col_obj.color().getRgb()
                size = font.pointSize()
                bold = font.bold()
                row_index = row_tmp + i - koef_hide
                column_index = column_tmp + j - koef_hide_c
                cell = sheet.cell(row=row_index, column=column_index)
                font_color = f"{r:02X}{g:02X}{b:02X}"
                cell.font = Font(color=font_color, size=size, bold=bold, name='Arial')
                val = tbl.item(i, j).text()
                r, g, b, a = item.background().color().getRgb()
                if F.is_numeric(val):
                    if isinstance(F.valm(val), int):
                        try:
                            sheet.cell(row=row_index, column=column_index).number_format = '0'
                        except:
                            pass
                    else:
                        try:
                            sheet.cell(row=row_index, column=column_index).number_format = '0.00'
                        except:
                            pass
                else:
                    sheet.cell(row=row_index, column=column_index).number_format = '@'
                cell.value = val
                cell.font = Font(color=font_color, size=size, bold=bold, name='Arial')
                cell.alignment = Alignment(wrap_text=True)
                if r == g == b == 0:
                    pass
                else:
                    fill_color = f"{r:02X}{g:02X}{b:02X}"
                    sheet.cell(row=row_index, column=column_index).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    row_tmp = copy.deepcopy(row) + 1
    column_tmp = copy.deepcopy(column) + 1
    if print_hat_tbl == False:
        row_tmp -= 1
    if tbl.verticalHeaderItem(0) == None:
        column_tmp -= 1

    def set_border(rng, type, weigth, color):
        rng.api.Borders(type).Weight = weigth + 1
        rng.api.Borders(type).Color = color

    # rng = sheet.range((row_tmp, column_tmp), (row_tmp + row_counter - 1, column_tmp + column_counter - 1))

    if 'custBorderInfo' in tbl.__dict__:
        custBorderInfo = tbl.custBorderInfo
        for i in range(row_tmp - 1, row_tmp + row_counter - 1):
            for j in range(column_tmp - 1, column_tmp + column_counter - 1):
                cell = sheet.cell(row=i + 1, column=j + 1)
                border = Border()
                out_ = "{:02X}{:02X}{:02X}{:02X}".format(
                    custBorderInfo.color_out.alpha(), custBorderInfo.color_out.red(), custBorderInfo.color_out.green(), custBorderInfo.color_out.blue()
                )
                in_ = "{:02X}{:02X}{:02X}{:02X}".format(
                    custBorderInfo.color_in.alpha(), custBorderInfo.color_in.red(), custBorderInfo.color_in.green(), custBorderInfo.color_in.blue()
                )

                if custBorderInfo.thick_out > 0:
                    if (i, j) in custBorderInfo.filled_bottom:
                        border.bottom = Side(style='thick', color=out_)
                    if (i, j) in custBorderInfo.filled_left:
                        border.left = Side(style='thick', color=out_)
                    if (i, j) in custBorderInfo.filled_right:
                        border.right = Side(style='thick', color=out_)
                    if (i, j) in custBorderInfo.filled_top:
                        border.top = Side(style='thick', color=out_)

                if custBorderInfo.thick_in > 0:
                    if (i, j) in custBorderInfo.inside_right:
                        border.right = Side(style='thin', color=in_)
                    if (i, j) in custBorderInfo.inside_top:
                        border.top = Side(style='thin', color=in_)
                cell.border = border

    else:
        rng = sheet.iter_rows(min_row=row_tmp, max_row=row_tmp + row_counter - 1,
                              min_col=column_tmp, max_col=column_tmp + column_counter - 1)

        def format_borders(rng, weight, XlBordersIndex='All', color_str: str = '0,0,0'):
            r, g, b = [int(_) for _ in color_str.split(',')]
            color = f"{r:02X}{g:02X}{b:02X}"  # Преобразуем RGB в HEX

            border_style = Side(style='thin' if weight == 1 else 'thick', color=color)

            try:
                for row in rng:
                    for cell in row:
                        if XlBordersIndex == 'xlCrossoutAll':
                            # Применяем границы ко всем сторонам
                            cell.border = Border(left=border_style, right=border_style, top=border_style,
                                                 bottom=border_style)
                        elif XlBordersIndex == 'xlBottomRightAll':
                            # Применяем границы только к нижней и правой сторонам
                            cell.border = Border(right=border_style, bottom=border_style)
                        else:
                            # Применяем границы в зависимости от XlBordersIndex
                            if XlBordersIndex == 'All':
                                cell.border = Border(left=border_style, right=border_style, top=border_style,
                                                     bottom=border_style)
                            else:
                                # Здесь можно добавить обработку других границ, если нужно
                                return f'XlBordersIndex = "{XlBordersIndex}" not found. Formatted all edges.'
            except Exception as e:
                return f'Exception = {e}'

        format_borders(rng, 1, color_str='133,133,133')
    workbook.save(file_path)

    rez = file_path
    return rez



def format_borders(xl_range_obj, weight, XlBordersIndex='All', color_str:str = '0,0,0'):
    color = XL.utils.rgb_to_int([int(_) for _ in color_str.split(',')])
    # documentation for borders object:
    #   https://learn.microsoft.com/en-us/office/vba/api/excel.borders
    # documentation for border object:
    #   https://learn.microsoft.com/en-us/office/vba/api/excel.border(object)
    # enumeration for xlbordersindex object:
    #   https://learn.microsoft.com/en-us/office/vba/api/excel.xlbordersindex
    # enumeration for xlborderweight object
    # https://learn.microsoft.com/en-us/office/vba/api/excel.xlborderweight

    try:


        # Custom function to "cross out" all cells in the range
        if XlBordersIndex == 'xlCrossoutAll':
            xl_range_obj.api.Borders(5).Weight = weight
            xl_range_obj.api.Borders(5).Color = color
            xl_range_obj.api.Borders(6).Weight = weight
            xl_range_obj.api.Borders(6).Color = color
            return ''

        # Custom function to format the bottom and right edges for all cells in the range
        elif XlBordersIndex == 'xlBottomRightAll':
            xl_range_obj.api.Borders(2).Weight = weight
            xl_range_obj.api.Borders(2).Color = color
            xl_range_obj.api.Borders(4).Weight = weight
            xl_range_obj.api.Borders(4).Color = color
            return ''

        else:
            edge = BORDER_DICT.get(XlBordersIndex, 0)
            if edge:
                xl_range_obj.api.Borders(edge).Weight = weight
                xl_range_obj.api.Borders(edge).Color = color
                return ''
            else:
                xl_range_obj.api.Borders.Weight = weight
                xl_range_obj.api.Borders.Color = color
                if XlBordersIndex != 'All':
                    return f'XlBordersIndex = "{XlBordersIndex}" not found.  Formatted all edges.'
                else:
                    return ''

    except Exception as e:
        return f'Exception = {e}'


def px2ch(px):
    
    """there are two magic
    numbers: 1 / 12 and 1 / 7 that
    found experimentally // convertion
    assumes that Normal style has
    default font of "Calibri 11pt" // and display is 96
    ppi // !!!note: px  must be integer"""
    px = int(px * 0.876)
    if px == 0:
        return 0
    elif px < 12:
        return round(px / 12 * 100) / 100

    else:
        return round((1 + (px - 12) / 7) * 100) / 100


class ExcelParser:
    def __init__(self, file_path: str):
        self.work_book = xlo.load_workbook(filename=file_path, data_only=True)

    @property
    def worksheets(self) -> list[str]:
        """Список названий листов инициализированной книги"""
        return [obj.title for obj in self.work_book.worksheets]

    def make_row_generator(
            self,
            sheet_name: str,
            min_row: int = None,
            max_col: int = None
    ) -> Generator[tuple[str], None, None]:
        """
            Возвращает генератор строчек excel
        """
        if sheet_name not in self.worksheets:
            return
        sheet = self.work_book[sheet_name]
        for row in sheet.iter_rows(values_only=True, min_row=min_row, max_row=max_col):
            yield row

    def find_widest_row(self, data_list):
        widest_row = []
        row_idx = None
        max_width = 0
        for idx, row in enumerate(data_list):
            current_width = len([cell for cell in row if cell is not None])
            if current_width > max_width:
                max_width = current_width
                widest_row = row
                row_idx = idx
        return widest_row, row_idx

    def data_by_worksheet(self, cur_sheet: str):
        row_data = self.make_row_generator(cur_sheet)
        row_idx = None
        max_width = 0
        result = []
        for idx, row in enumerate(row_data):
            result.append([self.transform_value(col) for col in row])
            current_width = len([cell for cell in row if cell is not None])
            if current_width > max_width:
                max_width = current_width
                row_idx = idx
        if row_idx is not None:
            header = result[row_idx]

            return [dict(zip(header, row)) for row in result[row_idx + 1:]]
        return result

    def transform_value(self, val: Any):
        if val is None:
            return ''
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d %H:%M:%S')
        return str(val)


def prepare_docx_data(data, bold_header: bool):
    if len(data) == 0:
        return
    if isinstance(data[0], dict):
        list_of_lists = F.list_of_dicts_to_list_of_lists(data)
    list_of_lists[0] = [
        RichText(head, bold=bold_header)
        for head in data[0]
    ]
    return list_of_lists


def make_docx_report(
        # Данные
        report_name: str,
        input_rows: list[dict],
        output_rows: list[dict],
        input_capture: str = 'Исходные данные',
        output_capture: str = 'Расчётные данные',
        # Настройки оформления
        bold_input_headers: bool = True,
        bold_input_capture: bool = True,
        bold_output_headers: bool = True,
        bold_output_capture: bool = True,
        # Пути используемых файлов
        template_name: str = "report.docx",
        output_docx_path: str = "output.docx"
):
    doc = DocxTemplate(template_name)
    output_data = prepare_docx_data(output_rows, bold_output_headers)
    input_data = prepare_docx_data(input_rows, bold_input_headers)
    data = {
        'report_name': report_name,
        'input': input_data,
        'i_capture': RichText(input_capture, bold=bold_input_capture),
        'i_len': len(input_data[0]),
        'output': output_data,
        'o_capture': RichText(output_capture, bold=bold_output_capture),
        'o_len': len(output_data[0]),
    }
    doc.render(data)
    doc.save(output_docx_path)
    return output_docx_path